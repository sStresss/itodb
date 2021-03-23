# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'PanelJournal_v15.ui'
#
# Created by: PyQt5 UI code generator 5.14.2
#
# WARNING! All changes made in this file will be lost!


from PyQt5 import QtGui
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt
from AddNewOStuff import Ui_AddNewOStuff
from AddNewKStuff import Ui_AddNewKStuff
import GlobalValues
import time
import sys
import threading
import pymysql
import easygui
from shutil import copy
from SystemSettings import Ui_SysSettings
from PanelMesBox import Ui_MesBox
from PanelOTransfer import Ui_OTransfer
from PanelKTransfer import Ui_KTransfer
from PanelEditOStuff import Ui_EditOStuff
from PanelEditKStuff import Ui_EditKStuff
from Journal import Ui_Journal
from PanelDelReason import Ui_DelReason
from PanelStat import Ui_Stat
from PanelGlobalStat import Ui_GlobalStat
from PanelDialogBox import Ui_DialogBox
from PanelInfo import Ui_Info
from PanelOGroupTransfer import Ui_OGroupTransfer
from PanelKGroupTransfer import Ui_KGroupTransfer
from PanelOGroupEdit import Ui_OGroupEdit
from PanelKGroupEdit import Ui_KGroupEdit
import subprocess
import os
import socket
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


PDFJS = 'file:///C:/itoDB/webEngine/web/viewer.html'

def oSpecFun():
    subb = subprocess.Popen(['C:/itoDB/pdfviewer/pdfviewer.exe'],shell=True)
    th = threading.Thread(target=tcpclient)
    th.start()

def tcpclient():
    data = ''
    tryConn = False
    HOST = '127.0.0.1'  # Standard loopback interface address (localhost)
    PORT = 65101  # Port to listen on (non-privileged ports are > 1023)

    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:

        while data == '':
            try:
                s.connect((HOST, PORT))
                data = str(GlobalValues.TreeParentName)
                data = data.encode()
                s.sendall(data)
                data = s.recv(1024)
                print('connection succesful!')
                time.sleep(0.2)
                tryConn = True
            except:
                tryConn = False
                print("waiting connection....")

    # print('Received', repr(data))

class Ui_MainDB(QDialog):

    def __init__(self):
        super().__init__()
        self.runUi()

    def runUi(self):
        self.setObjectName("Dialog")
        self.resize(1920, 1011)
        self.setWindowOpacity(1.0)
        self.MainFrame = QtWidgets.QFrame(self)
        self.MainFrame.setEnabled(True)
        self.MainFrame.setGeometry(QtCore.QRect(0, 0, 1920, 1011))
        self.MainFrame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.MainFrame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.MainFrame.setObjectName("MainFrame")
        self.btn_RefreshOTbl = QtWidgets.QPushButton(self.MainFrame)
        self.btn_RefreshOTbl.setGeometry(QtCore.QRect(121, 9, 25, 25))
        self.btn_RefreshOTbl.setStyleSheet("")
        self.btn_RefreshOTbl.setObjectName("btn_RefreshOTbl")
        self.le_OStuff = QtWidgets.QLineEdit(self.MainFrame)
        self.le_OStuff.setGeometry(QtCore.QRect(1400, 12, 311, 24))
        self.le_OStuff.setText("")
        self.le_OStuff.setObjectName("le_OStuff")
        self.cb_OStuff = QtWidgets.QComboBox(self.MainFrame)
        self.cb_OStuff.setGeometry(QtCore.QRect(1720, 12, 111, 24))
        self.cb_OStuff.setObjectName("cb_OStuff")
        self.cb_OStuff.addItem("")
        self.cb_OStuff.addItem("")
        self.cb_OStuff.addItem("")
        self.btn_SrchKStuff = QtWidgets.QPushButton(self.MainFrame)
        self.btn_SrchKStuff.setGeometry(QtCore.QRect(1840, 512, 61, 24))
        self.btn_SrchKStuff.setObjectName("btn_SrchKStuff")
        self.tree_Obj = QtWidgets.QTreeWidget(self.MainFrame)
        self.tree_Obj.setGeometry(QtCore.QRect(10, 10, 157, 991))
        self.tree_Obj.setObjectName("tree_Obj")
        self.cb_KStuff = QtWidgets.QComboBox(self.MainFrame)
        self.cb_KStuff.setGeometry(QtCore.QRect(1720, 512, 111, 24))
        self.cb_KStuff.setObjectName("cb_KStuff")
        self.cb_KStuff.addItem("")
        self.cb_KStuff.addItem("")
        self.cb_KStuff.addItem("")
        self.tbl_OStuff = QtWidgets.QTableWidget(self.MainFrame)
        self.tbl_OStuff.setGeometry(QtCore.QRect(171, 50, 1731, 451))
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.tbl_OStuff.sizePolicy().hasHeightForWidth())
        self.tbl_OStuff.setSizePolicy(sizePolicy)
        self.tbl_OStuff.setContextMenuPolicy(QtCore.Qt.ActionsContextMenu)
        self.tbl_OStuff.setStyleSheet("QTableWidget {background-color: rgb(255,255,255);\n"
                                      "border: 1px solid rgb(150,150,150);\n"
                                      "gridline-color: rgb(42,42,42);\n"
                                      "border-radius:0px;\n"
                                      "border-bottom-right-radius: 0px;\n"
                                      "color:black;}\n"
                                      "\n"
                                      "QLineEdit {background-color: white;}\n"
                                      "\n"
                                      "QHeaderView::section {\n"
                                      "gridline-color: rgb(89,89,89);\n"
                                      "background-color: rgb(142,187,208);\n"
                                      "color: black;};")
        self.tbl_OStuff.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.tbl_OStuff.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.tbl_OStuff.setAutoScrollMargin(5)
        self.tbl_OStuff.setEditTriggers(
            QtWidgets.QAbstractItemView.AnyKeyPressed | QtWidgets.QAbstractItemView.DoubleClicked | QtWidgets.QAbstractItemView.EditKeyPressed | QtWidgets.QAbstractItemView.SelectedClicked)
        self.tbl_OStuff.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.tbl_OStuff.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.tbl_OStuff.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerItem)
        self.tbl_OStuff.setHorizontalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        self.tbl_OStuff.setObjectName("tbl_OStuff")
        self.tbl_OStuff.setColumnCount(11)
        self.tbl_OStuff.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_OStuff.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_OStuff.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_OStuff.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_OStuff.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_OStuff.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_OStuff.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_OStuff.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_OStuff.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_OStuff.setHorizontalHeaderItem(8, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_OStuff.setHorizontalHeaderItem(9, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_OStuff.setHorizontalHeaderItem(10, item)
        self.tbl_OStuff.horizontalHeader().setCascadingSectionResizes(False)
        self.tbl_OStuff.horizontalHeader().setDefaultSectionSize(157)
        self.tbl_OStuff.horizontalHeader().setHighlightSections(True)
        self.tbl_OStuff.verticalHeader().setVisible(False)
        self.btn_SrchOStuff = QtWidgets.QPushButton(self.MainFrame)
        self.btn_SrchOStuff.setGeometry(QtCore.QRect(1840, 12, 61, 24))
        self.btn_SrchOStuff.setObjectName("btn_SrchOStuff")
        self.btn_AddKStuff = QtWidgets.QPushButton(self.MainFrame)
        self.btn_AddKStuff.setGeometry(QtCore.QRect(173, 512, 93, 24))
        self.btn_AddKStuff.setStyleSheet("border-radius:0px;\n"
                                         "font: 9pt \"MS Shell Dlg 2\";")
        self.btn_AddKStuff.setObjectName("btn_AddKStuff")
        self.tbl_KStuff = QtWidgets.QTableWidget(self.MainFrame)
        self.tbl_KStuff.setGeometry(QtCore.QRect(171, 550, 1731, 451))
        self.tbl_KStuff.setStyleSheet("QTableWidget {background-color: rgb(255,255,255);\n"
                                      "border: 1px solid rgb(150,150,150);\n"
                                      "gridline-color: rgb(42,42,42);\n"
                                      "border-radius:0px;\n"
                                      "border-bottom-right-radius: 0px;\n"
                                      "color:black;}\n"
                                      "\n"
                                      "QLineEdit {background-color: white;}\n"
                                      "\n"
                                      "QHeaderView::section {\n"
                                      "gridline-color: rgb(89,89,89);\n"
                                      "background-color: rgb(142,187,208);\n"
                                      "color: black;};")
        self.tbl_KStuff.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.tbl_KStuff.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.tbl_KStuff.setEditTriggers(
            QtWidgets.QAbstractItemView.AnyKeyPressed | QtWidgets.QAbstractItemView.DoubleClicked | QtWidgets.QAbstractItemView.EditKeyPressed | QtWidgets.QAbstractItemView.SelectedClicked)
        self.tbl_KStuff.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        self.tbl_KStuff.setObjectName("tbl_KStuff")
        self.tbl_KStuff.setColumnCount(11)
        self.tbl_KStuff.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_KStuff.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_KStuff.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_KStuff.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_KStuff.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_KStuff.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_KStuff.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_KStuff.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_KStuff.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_KStuff.setHorizontalHeaderItem(8, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_KStuff.setHorizontalHeaderItem(9, item)
        item = QtWidgets.QTableWidgetItem()
        self.tbl_KStuff.setHorizontalHeaderItem(10, item)
        self.tbl_KStuff.horizontalHeader().setDefaultSectionSize(157)
        self.le_KStuff = QtWidgets.QLineEdit(self.MainFrame)
        self.le_KStuff.setGeometry(QtCore.QRect(1400, 512, 311, 22))
        self.le_KStuff.setObjectName("le_KStuff")
        self.btn_Settings = QtWidgets.QPushButton(self.MainFrame)
        self.btn_Settings.setGeometry(QtCore.QRect(1065, 12, 93, 24))
        self.btn_Settings.setObjectName("btn_Settings")
        self.btn_ExelTableData = QtWidgets.QPushButton(self.MainFrame)
        self.btn_ExelTableData.setGeometry(QtCore.QRect(960, 12, 93, 24))
        self.btn_ExelTableData.setObjectName("btn_ExelTableData")
        self.btn_AddOStuff = QtWidgets.QPushButton(self.MainFrame)
        self.btn_AddOStuff.setGeometry(QtCore.QRect(170, 15, 81, 16))
        self.btn_AddOStuff.setStyleSheet("border-radius:0px;\n"
                                         "font: 9pt \"MS Shell Dlg 2\";")
        self.btn_AddOStuff.setObjectName("btn_AddOStuff")
        self.vScrl_OTbl = QtWidgets.QScrollBar(self.MainFrame)
        self.vScrl_OTbl.setGeometry(QtCore.QRect(1900, 73, 11, 428))
        self.vScrl_OTbl.setStyleSheet("QScrollBar:vertical {\n"
                                      "border: 1px solid rgb(89,89,89);\n"
                                      " background: rgb(255,255,255);\n"
                                      "width:10px;\n"
                                      "margin: 0px 0px 0px 0px;\n"
                                      "}\n"
                                      "QScrollBar::handle:vertical {\n"
                                      "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0 rgb(142,187,208), stop: 0.5 rgb(142,187,208), stop:1 rgb(142,187,208));\n"
                                      "min-height: 5px;\n"
                                      "}\n"
                                      "QScrollBar::add-line:vertical {\n"
                                      " background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop: 0 rgb(142,187,208), stop: 0.5 rgb(142,187,208),  stop:1 rgb(142,187,208));\n"
                                      "height: 0px;\n"
                                      "subcontrol-position: bottom;\n"
                                      "subcontrol-origin: margin;\n"
                                      "}\n"
                                      "QScrollBar::sub-line:vertical {\n"
                                      "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0  rgb(142,187,208), stop: 0.5 rgb(142,187,208),  stop:1 rgb(142,187,208));\n"
                                      "height: 0 px;\n"
                                      "subcontrol-position: top;\n"
                                      "subcontrol-origin: margin;\n"
                                      "}")
        self.vScrl_OTbl.setMaximum(101)
        self.vScrl_OTbl.setOrientation(QtCore.Qt.Vertical)
        self.vScrl_OTbl.setObjectName("vScrl_OTbl")
        self.vScrl_KTbl_2 = QtWidgets.QScrollBar(self.MainFrame)
        self.vScrl_KTbl_2.setGeometry(QtCore.QRect(1900, 573, 11, 428))
        self.vScrl_KTbl_2.setStyleSheet("QScrollBar:vertical {\n"
                                        "border: 1px solid rgb(89,89,89);\n"
                                        " background: rgb(255,255,255);\n"
                                        "width:10px;\n"
                                        "margin: 0px 0px 0px 0px;\n"
                                        "}\n"
                                        "QScrollBar::handle:vertical {\n"
                                        "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0 rgb(142,187,208), stop: 0.5 rgb(142,187,208), stop:1 rgb(142,187,208));\n"
                                        "min-height: 5px;\n"
                                        "}\n"
                                        "QScrollBar::add-line:vertical {\n"
                                        " background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop: 0 rgb(142,187,208), stop: 0.5 rgb(142,187,208),  stop:1 rgb(142,187,208));\n"
                                        "height: 0px;\n"
                                        "subcontrol-position: bottom;\n"
                                        "subcontrol-origin: margin;\n"
                                        "}\n"
                                        "QScrollBar::sub-line:vertical {\n"
                                        "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0  rgb(142,187,208), stop: 0.5 rgb(142,187,208),  stop:1 rgb(142,187,208));\n"
                                        "height: 0 px;\n"
                                        "subcontrol-position: top;\n"
                                        "subcontrol-origin: margin;\n"
                                        "}")
        self.vScrl_KTbl_2.setOrientation(QtCore.Qt.Vertical)
        self.vScrl_KTbl_2.setObjectName("vScrl_KTbl_2")
        self.btn_DelObj = QtWidgets.QPushButton(self.MainFrame)
        self.btn_DelObj.setGeometry(QtCore.QRect(143, 53, 25, 25))
        self.btn_DelObj.setObjectName("btn_DelObj")
        self.btn_AddObj = QtWidgets.QPushButton(self.MainFrame)
        self.btn_AddObj.setGeometry(QtCore.QRect(143, 9, 25, 25))
        self.btn_AddObj.setObjectName("btn_AddObj")
        self.line = QtWidgets.QFrame(self.MainFrame)
        self.line.setGeometry(QtCore.QRect(11, 32, 149, 3))
        self.line.setFrameShape(QtWidgets.QFrame.HLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.btn_AddChildObj = QtWidgets.QPushButton(self.MainFrame)
        self.btn_AddChildObj.setGeometry(QtCore.QRect(143, 31, 25, 25))
        self.btn_AddChildObj.setObjectName("btn_AddChildObj")
        self.label_3 = QtWidgets.QLabel(self.MainFrame)
        self.label_3.setGeometry(QtCore.QRect(1180, 12, 131, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.le_OViewMode = QtWidgets.QLabel(self.MainFrame)
        self.le_OViewMode.setGeometry(QtCore.QRect(1310, 12, 81, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.le_OViewMode.setFont(font)
        self.le_OViewMode.setStyleSheet("color: rgb(0,0,0);")
        self.le_OViewMode.setText("")
        self.le_OViewMode.setAlignment(QtCore.Qt.AlignCenter)
        self.le_OViewMode.setObjectName("le_OViewMode")
        self.le_KViewMode = QtWidgets.QLabel(self.MainFrame)
        self.le_KViewMode.setGeometry(QtCore.QRect(1310, 512, 81, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.le_KViewMode.setFont(font)
        self.le_KViewMode.setStyleSheet("color: rgb(0,0,0);")
        self.le_KViewMode.setText("")
        self.le_KViewMode.setAlignment(QtCore.Qt.AlignCenter)
        self.le_KViewMode.setObjectName("le_KViewMode")
        self.label_4 = QtWidgets.QLabel(self.MainFrame)
        self.label_4.setGeometry(QtCore.QRect(1180, 512, 131, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.line_2 = QtWidgets.QFrame(self.MainFrame)
        self.line_2.setGeometry(QtCore.QRect(1170, 35, 220, 3))
        self.line_2.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")
        self.line_3 = QtWidgets.QFrame(self.MainFrame)
        self.line_3.setGeometry(QtCore.QRect(1170, 536, 220, 3))
        self.line_3.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_3.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_3.setObjectName("line_3")
        self.chB_OGroup = QtWidgets.QCheckBox(self.MainFrame)
        self.chB_OGroup.setGeometry(QtCore.QRect(175, 53, 16, 17))
        self.chB_OGroup.setText("")
        self.chB_OGroup.setObjectName("chB_OGroup")
        self.chB_KGroup = QtWidgets.QCheckBox(self.MainFrame)
        self.chB_KGroup.setGeometry(QtCore.QRect(175, 554, 16, 17))
        self.chB_KGroup.setText("")
        self.chB_KGroup.setObjectName("chB_KGroup")
        self.btn_Journal = QtWidgets.QPushButton(self.MainFrame)
        self.btn_Journal.setGeometry(QtCore.QRect(845, 12, 101, 24))
        self.btn_Journal.setObjectName("btn_Journal")
        self.line_7 = QtWidgets.QFrame(self.MainFrame)
        self.line_7.setGeometry(QtCore.QRect(270, 512, 3, 24))
        self.line_7.setFrameShape(QtWidgets.QFrame.VLine)
        self.line_7.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_7.setObjectName("line_7")
        self.btn_Spec = QtWidgets.QPushButton(self.MainFrame)
        self.btn_Spec.setGeometry(QtCore.QRect(455, 12, 91, 24))
        self.btn_Spec.setObjectName("btn_Spec")
        self.line_11 = QtWidgets.QFrame(self.MainFrame)
        self.line_11.setGeometry(QtCore.QRect(1390, 12, 3, 24))
        self.line_11.setFrameShape(QtWidgets.QFrame.VLine)
        self.line_11.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_11.setObjectName("line_11")
        self.line_12 = QtWidgets.QFrame(self.MainFrame)
        self.line_12.setGeometry(QtCore.QRect(1170, 12, 3, 24))
        self.line_12.setFrameShape(QtWidgets.QFrame.VLine)
        self.line_12.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_12.setObjectName("line_12")
        self.line_13 = QtWidgets.QFrame(self.MainFrame)
        self.line_13.setGeometry(QtCore.QRect(1170, 512, 3, 24))
        self.line_13.setFrameShape(QtWidgets.QFrame.VLine)
        self.line_13.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_13.setObjectName("line_13")
        self.line_14 = QtWidgets.QFrame(self.MainFrame)
        self.line_14.setGeometry(QtCore.QRect(1390, 512, 3, 24))
        self.line_14.setFrameShape(QtWidgets.QFrame.VLine)
        self.line_14.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_14.setObjectName("line_14")
        self.btn_PhotoArchive = QtWidgets.QPushButton(self.MainFrame)
        self.btn_PhotoArchive.setGeometry(QtCore.QRect(555, 12, 91, 24))
        self.btn_PhotoArchive.setObjectName("btn_PhotoArchive")
        self.btn_Documentation = QtWidgets.QPushButton(self.MainFrame)
        self.btn_Documentation.setGeometry(QtCore.QRect(355, 12, 91, 24))
        self.btn_Documentation.setObjectName("btn_Documentation")
        self.btn_TransportDocs = QtWidgets.QPushButton(self.MainFrame)
        self.btn_TransportDocs.setGeometry(QtCore.QRect(655, 12, 91, 24))
        self.btn_TransportDocs.setObjectName("btn_TransportDocs")
        self.btn_StatOTbl = QtWidgets.QPushButton(self.MainFrame)
        self.btn_StatOTbl.setGeometry(QtCore.QRect(170, 36, 16, 16))
        self.btn_StatOTbl.setStyleSheet("")
        self.btn_StatOTbl.setObjectName("btn_StatOTbl")
        self.btn_StatKTbl = QtWidgets.QPushButton(self.MainFrame)
        self.btn_StatKTbl.setGeometry(QtCore.QRect(170, 536, 16, 16))
        self.btn_StatKTbl.setObjectName("btn_StatKTbl")
        self.btn_NetworkConfig = QtWidgets.QPushButton(self.MainFrame)
        self.btn_NetworkConfig.setGeometry(QtCore.QRect(755, 12, 71, 24))
        self.btn_NetworkConfig.setObjectName("btn_NetworkConfig")
        self.btn_Statistic = QtWidgets.QPushButton(self.MainFrame)
        self.btn_Statistic.setGeometry(QtCore.QRect(265, 12, 81, 24))
        self.btn_Statistic.setObjectName("btn_Statistic")
        self.line_8 = QtWidgets.QFrame(self.MainFrame)
        self.line_8.setGeometry(QtCore.QRect(255, 12, 3, 24))
        self.line_8.setFrameShape(QtWidgets.QFrame.VLine)
        self.line_8.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_8.setObjectName("line_8")
        self.line_9 = QtWidgets.QFrame(self.MainFrame)
        self.line_9.setGeometry(QtCore.QRect(835, 12, 3, 24))
        self.line_9.setFrameShape(QtWidgets.QFrame.VLine)
        self.line_9.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_9.setObjectName("line_9")
        self.btn_StatOTbl.raise_()
        self.le_OStuff.raise_()
        self.cb_OStuff.raise_()
        self.btn_SrchKStuff.raise_()
        self.tree_Obj.raise_()
        self.cb_KStuff.raise_()
        self.tbl_OStuff.raise_()
        self.btn_SrchOStuff.raise_()
        self.btn_AddKStuff.raise_()
        self.tbl_KStuff.raise_()
        self.le_KStuff.raise_()
        self.btn_Settings.raise_()
        self.btn_ExelTableData.raise_()
        self.btn_AddOStuff.raise_()
        self.vScrl_OTbl.raise_()
        self.vScrl_KTbl_2.raise_()
        self.btn_DelObj.raise_()
        self.btn_AddObj.raise_()
        self.line.raise_()
        self.btn_AddChildObj.raise_()
        self.label_3.raise_()
        self.le_OViewMode.raise_()
        self.le_KViewMode.raise_()
        self.label_4.raise_()
        self.line_2.raise_()
        self.line_3.raise_()
        self.chB_OGroup.raise_()
        self.chB_KGroup.raise_()
        self.btn_Journal.raise_()
        self.line_7.raise_()
        self.btn_Spec.raise_()
        self.line_11.raise_()
        self.line_12.raise_()
        self.line_13.raise_()
        self.line_14.raise_()
        self.btn_PhotoArchive.raise_()
        self.btn_Documentation.raise_()
        self.btn_TransportDocs.raise_()
        self.btn_StatKTbl.raise_()
        self.btn_NetworkConfig.raise_()
        self.btn_Statistic.raise_()
        self.btn_RefreshOTbl.raise_()
        self.line_8.raise_()
        self.line_9.raise_()

        #контекст для таблиц
        self.tbl_OStuff.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.tbl_OStuff.customContextMenuRequested.connect(self.generateMenu)
        self.tbl_OStuff.viewport().installEventFilter(self)
        self.layout = QVBoxLayout()
        self.layout.addWidget(self.tbl_OStuff)

        self.tbl_KStuff.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.tbl_KStuff.customContextMenuRequested.connect(self.generateMenu)
        self.tbl_KStuff.viewport().installEventFilter(self)
        self.layout2 = QVBoxLayout()
        self.layout2.addWidget(self.tbl_KStuff)

        #контекст для дерева обьектов
        self.tree_Obj.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.tree_Obj.customContextMenuRequested.connect(self.generateTreeMenu)
        self.tree_Obj.viewport().installEventFilter(self)

        self.menuTree = QMenu(self)
        self.menuTree.setStyleSheet("""
                              QMenu {border: 1px inset grey; background-color: #fff; color: #000; padding: -1; padding-left: -6;}
                              QMenu:selected {background-color: #ddf; color: #000;}
                          """)
        self.menuTree.setFixedWidth(110)

        self.menu_actionTree = QAction('Переименовать', self)
        self.menu_actionTree.setData('option1')
        self.menu_actionTree.triggered.connect(self.renameTreeObj)
        self.menuTree.addAction(self.menu_actionTree)

        self.menu_actionInfo = QAction('Заметки', self)
        self.menu_actionInfo.setData('option2')
        self.menu_actionInfo.triggered.connect(self.openInfo)
        self.menuTree.addAction(self.menu_actionInfo)

        self.menu_actionExel = QAction('Выгрузить в Exel', self)
        self.menu_actionExel.setData('option3')
        self.menu_actionExel.triggered.connect(self.makeExel)
        self.menuTree.addAction(self.menu_actionExel)
        #контекст для кнопки спецификации
        self.btn_Spec.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.btn_Spec.customContextMenuRequested.connect(self.generateBtnMenu)

        self.menuSpec = QMenu(self)
        self.menuSpec.setStyleSheet("""
                            QMenu {border: 1px inset grey; background-color: #fff; color: #000; padding: -1; padding-left: -6;}
                            QMenu:selected {background-color: #ddf; color: #000;}
                        """)
        self.menuSpec.setFixedWidth(80)

        self.menu_actionBtnSpec = QAction('Загрузить', self)
        self.menu_actionBtnSpec.setData('option1')
        self.menu_actionBtnSpec.triggered.connect(self.reloadSpec)
        self.menuSpec.addAction(self.menu_actionBtnSpec)

        #кнопка обновления состояния дереве
        self.btn_RefreshTreeStates = QtWidgets.QPushButton(self.MainFrame)
        self.btn_RefreshTreeStates.setGeometry(QtCore.QRect(99, 9, 25, 25))
        self.btn_RefreshTreeStates.setStyleSheet("")
        self.btn_RefreshTreeStates.setObjectName("btn_RefreshOTbl")

        #############################

        self.retranslateUi()
        QtCore.QMetaObject.connectSlotsByName(self)
        self.firstSets()
        self.treeUpdate()

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("Dialog", "itoDB_v0.28"))
        self.btn_RefreshOTbl.setText(_translate("Dialog", "^"))
        self.cb_OStuff.setItemText(0, _translate("Dialog", "Серийный номер"))
        self.cb_OStuff.setItemText(1, _translate("Dialog", "Цел.Объект"))
        self.cb_OStuff.setItemText(2, _translate("Dialog", "Поставщик"))
        self.btn_SrchKStuff.setText(_translate("Dialog", "Поиск"))
        self.tree_Obj.headerItem().setText(0, _translate("Dialog", "Объекты"))
        self.cb_KStuff.setItemText(0, _translate("Dialog", "Серийный номер"))
        self.cb_KStuff.setItemText(1, _translate("Dialog", "Цел.Объект"))
        self.cb_KStuff.setItemText(2, _translate("Dialog", "Поставщик"))
        self.tbl_OStuff.setSortingEnabled(False)
        item = self.tbl_OStuff.horizontalHeaderItem(0)
        item.setText(_translate("Dialog", "ID"))
        item = self.tbl_OStuff.horizontalHeaderItem(1)
        item.setText(_translate("Dialog", "Тип"))
        item = self.tbl_OStuff.horizontalHeaderItem(2)
        item.setText(_translate("Dialog", "Модель"))
        item = self.tbl_OStuff.horizontalHeaderItem(3)
        item.setText(_translate("Dialog", "Серийный номер"))
        item = self.tbl_OStuff.horizontalHeaderItem(4)
        item.setText(_translate("Dialog", "Производитель"))
        item = self.tbl_OStuff.horizontalHeaderItem(5)
        item.setText(_translate("Dialog", "Поставщик"))
        item = self.tbl_OStuff.horizontalHeaderItem(6)
        item.setText(_translate("Dialog", "Дата поступления"))
        item = self.tbl_OStuff.horizontalHeaderItem(7)
        item.setText(_translate("Dialog", "Целевой объект"))
        item = self.tbl_OStuff.horizontalHeaderItem(8)
        item.setText(_translate("Dialog", "Факт. Объект"))
        item = self.tbl_OStuff.horizontalHeaderItem(9)
        item.setText(_translate("Dialog", "Дата перемещения"))
        item = self.tbl_OStuff.horizontalHeaderItem(10)
        item.setText(_translate("Dialog", "Комментарий"))
        self.btn_SrchOStuff.setText(_translate("Dialog", "Поиск"))
        self.btn_AddKStuff.setText(_translate("Dialog", "Комплектующие"))
        item = self.tbl_KStuff.horizontalHeaderItem(0)
        item.setText(_translate("Dialog", "ID"))
        item = self.tbl_KStuff.horizontalHeaderItem(1)
        item.setText(_translate("Dialog", "Тип"))
        item = self.tbl_KStuff.horizontalHeaderItem(2)
        item.setText(_translate("Dialog", "Модель"))
        item = self.tbl_KStuff.horizontalHeaderItem(3)
        item.setText(_translate("Dialog", "Серийный номер"))
        item = self.tbl_KStuff.horizontalHeaderItem(4)
        item.setText(_translate("Dialog", "Производитель"))
        item = self.tbl_KStuff.horizontalHeaderItem(5)
        item.setText(_translate("Dialog", "Поставщик"))
        item = self.tbl_KStuff.horizontalHeaderItem(6)
        item.setText(_translate("Dialog", "Дата поступления"))
        item = self.tbl_KStuff.horizontalHeaderItem(7)
        item.setText(_translate("Dialog", "Целевой объект"))
        item = self.tbl_KStuff.horizontalHeaderItem(8)
        item.setText(_translate("Dialog", "Факт. объект"))
        item = self.tbl_KStuff.horizontalHeaderItem(9)
        item.setText(_translate("Dialog", "Дата перемещения"))
        item = self.tbl_KStuff.horizontalHeaderItem(10)
        item.setText(_translate("Dialog", "Комментарий"))
        self.btn_Settings.setText(_translate("Dialog", "Настройки"))
        self.btn_ExelTableData.setText(_translate("Dialog", "Таблица Учета"))
        self.btn_AddOStuff.setText(_translate("Dialog", "Оборудовние"))
        self.btn_DelObj.setText(_translate("Dialog", "-"))
        self.btn_AddObj.setText(_translate("Dialog", "+"))
        self.btn_AddChildObj.setText(_translate("Dialog", "++"))
        self.btn_RefreshTreeStates.setText(_translate("Dialog", "#"))
        self.label_3.setText(_translate("Dialog", "Режим отображения:"))
        self.label_4.setText(_translate("Dialog", "Режим отображения:"))
        self.btn_Journal.setText(_translate("Dialog", "Журнал событий"))
        self.btn_Spec.setText(_translate("Dialog", "Спецификация"))
        self.btn_PhotoArchive.setText(_translate("Dialog", "Фотоархив"))
        self.btn_Documentation.setText(_translate("Dialog", "Документация"))
        self.btn_TransportDocs.setText(_translate("Dialog", "Накладные"))
        self.btn_StatOTbl.setText(_translate("Dialog", "^"))
        self.btn_StatKTbl.setText(_translate("Dialog", "^"))
        self.btn_NetworkConfig.setText(_translate("Dialog", "Конфиги"))
        self.btn_Statistic.setText(_translate("Dialog", "Состояние"))

    def firstSets(self):

        self.setWindowFlags(QtCore.Qt.CustomizeWindowHint | QtCore.Qt.WindowCloseButtonHint | QtCore.Qt.WindowMinMaxButtonsHint )
        self.setWindowState(QtCore.Qt.WindowMaximized)

        #иконка
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("C:\itoDB\iconMainWindow.ico"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.setWindowIcon(icon)
        self.setWindowOpacity(1.0)
        #коннекты
        self.tbl_OStuff.setSelectionBehavior(QTableView.SelectRows)
        self.tbl_KStuff.setSelectionBehavior(QTableView.SelectRows)
        self.btn_AddOStuff.clicked.connect(self.openPanelAddNewOStuff)
        self.btn_AddKStuff.clicked.connect(self.openPanelAddNewKStuff)
        self.btn_Settings.clicked.connect(self.openPanelSystemSettings)
        self.vScrl_OTbl.valueChanged.connect(self.sync_func)
        # self.btn_DelOStuff.clicked.connect(self.DeleteORecord)
        self.tbl_OStuff.cellClicked.connect(self.cell_was_clicked)
        self.btn_Spec.clicked.connect(self.thOSPEC)
        self.btn_Documentation.clicked.connect(self.openDocumentation)
        self.btn_PhotoArchive.clicked.connect(self.openPhotoArchive)
        self.btn_TransportDocs.clicked.connect(self.openTranSource)
        self.btn_ExelTableData.clicked.connect(self.openExelDataTable)

        self.btn_AddObj.clicked.connect(self.openPanelAddNewObj)
        self.btn_DelObj.clicked.connect(self.DelTreeObj)
        self.btn_AddChildObj.clicked.connect(self.AddNewChildObject)
        # self.btn_moveOStuff.clicked.connect(self.openOTransfer)
        self.tree_Obj.clicked.connect(self.TreeClickEvent)
        self.btn_RefreshOTbl.clicked.connect(self.BackToMain)
        # self.btn_editOStuff.clicked.connect(self.openPanelEditOStuff)
        self.btn_SrchOStuff.clicked.connect(self.SearchOStuff)
        self.cb_OStuff.currentIndexChanged.connect(self.SearchEnable)
        self.chB_OGroup.clicked.connect(self.OTblMode)

        self.vScrl_KTbl_2.valueChanged.connect(self.sync_func2)
        # self.btn_DelKStuff.clicked.connect(self.DeleteKRecord)
        self.tbl_KStuff.cellClicked.connect(self.cell2_was_clicked)
        # self.btn_moveKStuff.clicked.connect(self.openKTransfer)
        # self.btn_editKStuff.clicked.connect(self.openPanelEditKStuff)
        self.btn_SrchKStuff.clicked.connect(self.SearchKStuff)
        self.chB_KGroup.clicked.connect(self.KTblMode)
        self.btn_Journal.clicked.connect(self.openPanelJournal)
        self.btn_StatOTbl.clicked.connect(self.getStatOTbl)
        self.btn_StatKTbl.clicked.connect(self.getStatKTbl)
        self.btn_Statistic.clicked.connect(self.openGlobalStat)
        self.btn_NetworkConfig.clicked.connect(self.openNetworkConf)

        self.btn_RefreshTreeStates.clicked.connect(self.thTreeColorStateUpdate)

        #подхват паслогов в sql при первом запуске
        passlog = open('C:\itoDB\sqlpasslog.txt', "r")
        l = [line.strip() for line in passlog]
        #print(l)
        GlobalValues.SqlHostname = str(l[0])
        GlobalValues.SqlPort = int(l[1])
        GlobalValues.SqlUserName = str(l[2])
        GlobalValues.SqlPwd = str(l[3])
        GlobalValues.globalSpecSource = str(l[4]) + 'specsource/'
        GlobalValues.localSpecSource = str(l[5]) + 'specsource/'
        GlobalValues.globalDocSource = str(l[4]) + 'docsource/'
        GlobalValues.globalPhotosource = str(l[4]) + 'photosource/'
        GlobalValues.globalTranSource = str(l[4]) + 'transource/'
        GlobalValues.globalNetConfSource = str(l[4]) + 'netconfsource/'
        GlobalValues.globalExelDataTableSourse = str(l[4]) + 'учет.xlsx'
        # GlobalValues.SqlHostname = "127.0.0.1"
        # GlobalValues.SqlPort = 3303
        # GlobalValues.SqlUserName = "itouser"
        # GlobalValues.SqlPwd = "sinaps281082"
        passlog.close()
        #проверка группы доступа пользователя
        if GlobalValues.AccesGroup != 'admin':
            self.btn_AddOStuff.setEnabled(False)
            self.btn_AddKStuff.setEnabled(False)
            self.btn_Settings.setEnabled(False)
            self.btn_AddObj.setEnabled(False)
            self.btn_AddChildObj.setEnabled(False)
            self.btn_DelObj.setEnabled(False)

        #прочая хрень
        self.btn_StatOTbl.setToolTip('Статистика')
        self.btn_StatKTbl.setToolTip('Статистика')
        GlobalValues.checkGlobalviewMode = True
        self.le_OViewMode.setText("Глобальный")
        self.le_KViewMode.setText("Глобальный")
        self.tbl_OStuff.verticalHeader().hide()
        self.tbl_OStuff.verticalScrollBar().hide()
        self.tbl_OStuff.horizontalScrollBar().hide()
        self.tbl_OStuff.verticalHeader().setDefaultSectionSize(4)
        self.tbl_KStuff.verticalHeader().hide()
        self.tbl_KStuff.verticalScrollBar().hide()
        self.tbl_KStuff.horizontalScrollBar().hide()
        self.tbl_KStuff.verticalHeader().setDefaultSectionSize(4)
        self.tbl_OStuff.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        self.tbl_KStuff.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        self.tbl_OStuff.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.tbl_KStuff.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)

        #ресайз ширины столбцов таблиц
        self.tbl_OStuff.setColumnWidth(0,60)
        self.tbl_OStuff.setColumnWidth(3,130)
        self.tbl_OStuff.setColumnWidth(4,100)
        self.tbl_OStuff.setColumnWidth(6,120)
        self.tbl_OStuff.setColumnWidth(7,150)
        self.tbl_OStuff.setColumnWidth(8,150)
        self.tbl_OStuff.setColumnWidth(9,125)
        self.tbl_OStuff.setColumnWidth(10,423)
        self.tbl_OStuff.horizontalHeader().setSectionResizeMode(0,QHeaderView.Fixed)
        self.tbl_OStuff.horizontalHeader().setSectionResizeMode(3,QHeaderView.Fixed)
        self.tbl_OStuff.horizontalHeader().setSectionResizeMode(4,QHeaderView.Fixed)
        self.tbl_OStuff.horizontalHeader().setSectionResizeMode(6,QHeaderView.Fixed)
        self.tbl_OStuff.horizontalHeader().setSectionResizeMode(7,QHeaderView.Fixed)
        self.tbl_OStuff.horizontalHeader().setSectionResizeMode(8,QHeaderView.Fixed)
        self.tbl_OStuff.horizontalHeader().setSectionResizeMode(9,QHeaderView.Fixed)
        self.tbl_OStuff.horizontalHeader().setSectionResizeMode(10,QHeaderView.Fixed)

        self.tbl_KStuff.setColumnWidth(0,60)
        self.tbl_KStuff.setColumnWidth(3,130)
        self.tbl_KStuff.setColumnWidth(4,100)
        self.tbl_KStuff.setColumnWidth(6,120)
        self.tbl_KStuff.setColumnWidth(7,150)
        self.tbl_KStuff.setColumnWidth(8,150)
        self.tbl_KStuff.setColumnWidth(9,125)
        self.tbl_KStuff.setColumnWidth(10,423)
        self.tbl_KStuff.horizontalHeader().setSectionResizeMode(0,QHeaderView.Fixed)
        self.tbl_KStuff.horizontalHeader().setSectionResizeMode(3,QHeaderView.Fixed)
        self.tbl_KStuff.horizontalHeader().setSectionResizeMode(4,QHeaderView.Fixed)
        self.tbl_KStuff.horizontalHeader().setSectionResizeMode(6,QHeaderView.Fixed)
        self.tbl_KStuff.horizontalHeader().setSectionResizeMode(7,QHeaderView.Fixed)
        self.tbl_KStuff.horizontalHeader().setSectionResizeMode(8,QHeaderView.Fixed)
        self.tbl_KStuff.horizontalHeader().setSectionResizeMode(9,QHeaderView.Fixed)
        self.tbl_KStuff.horizontalHeader().setSectionResizeMode(10,QHeaderView.Fixed)

        #градиенты для елементов дерева
        red = 255
        green = 255
        blue = 255
        back_color = QtGui.QColor(red, green, blue)
        self.tree_Obj.setStyleSheet('background-color: rgb(255, 255, 255);')
        # bColor = [80,209,224]

        self.green_gradient = QtGui.QLinearGradient(QtCore.QPointF(50, -20),
                                               QtCore.QPointF(80, 20))
        self.green_gradient.setColorAt(0.0, QtGui.QColor(0, 247, 123))
        self.green_gradient.setColorAt(1.0, back_color)

        self.yellow_gradient = QtGui.QLinearGradient(QtCore.QPointF(50, -20),
                                             QtCore.QPointF(80, 20))
        self.yellow_gradient.setColorAt(0.0, QtGui.QColor(255,255,0))
        self.yellow_gradient.setColorAt(1.0, back_color)

        self.red_gradient = QtGui.QLinearGradient(QtCore.QPointF(50, -20),
                                             QtCore.QPointF(80, 20))
        self.red_gradient.setColorAt(0.0, QtGui.QColor(255, 140, 142))
        self.red_gradient.setColorAt(1.0, back_color)

        # self.btnOTblRefresh.setDisabled(True)
        self.SqlConCheck()
        GlobalValues.checkThOTableUpdate = True
        GlobalValues.checkOTblUpdateEvent = True
        GlobalValues.SqlFirstConnectUpdate = True

        GlobalValues.checkThKTableUpdate = True
        GlobalValues.checkKTblUpdateEvent = True
        GlobalValues.SqlFirstConnectUpdate = True

        self.FirstOTblUpdate()
        self.FirstKTblUpdate()

        self.thUpdateOTable()
        self.btn_RefreshTreeStates.click()
        GlobalValues.checkTreeStatesUpdate = True

    #работа с оборудованием
    def openHystory(self):
        GlobalValues.checkTargetHystory = True
        self.btn_Journal.click()

    def openPanelAddNewOStuff(self):
        self.uiAddNewOStuff = Ui_AddNewOStuff()
        self.uiAddNewOStuff.exec_()

    def openPanelSystemSettings(self):
        self.uiSysSettings = Ui_SysSettings()
        self.uiSysSettings.exec_()

    def openOTransfer(self):

        if GlobalValues.checkOTblWasClicked == True :

            if GlobalValues.checkGroupOTbl == True:
                temp_lst_selected = []
                lst_selected = []
                selected = self.tbl_OStuff.selectedItems()
                for item in selected:

                    if item.column() == 0:
                        temp_lst_selected.append(item.data(0))
                    if item.column() == 1:
                        temp_lst_selected.append(item.data(0))
                    if item.column() == 3:
                        temp_lst_selected.append(item.data(0))
                    if item.column() == 8:
                        temp_lst_selected.append(item.data(0))
                        lst_selected.append(temp_lst_selected)
                        temp_lst_selected = []
                GlobalValues.otbl_selected_lst = lst_selected
                self.uiOGroupTransfer = Ui_OGroupTransfer()
                self.uiOGroupTransfer.exec_()
            else:
                self.uiOTransfer = Ui_OTransfer()
                self.uiOTransfer.exec_()
        else:
            GlobalValues.MesText = "Оборудование не выбрано!"
            self.openMesBox()

    def SqlConCheck(self):
        try:
            con = pymysql.connect(host=GlobalValues.SqlHostname,
                                  port=GlobalValues.SqlPort,
                                  user=GlobalValues.SqlUserName,
                                  passwd=GlobalValues.SqlPwd,
                                  db=GlobalValues.SqlDBName)
            with con:
                cur = con.cursor()
                cur.execute("SELECT VERSION()")

                version = cur.fetchone()

                if str(format(version[0])) != '':
                    GlobalValues.SqlConCheck = True

                #print("Database version: {}".format(version[0]))
            con.close()
        except:
            GlobalValues.SqlConCheck = False

    def thUpdateOTable(self):
        # print("Update Func start OK!!!")
        th =threading.Thread(target = self.OTableUpdate)
        th.start()

    def OTableUpdate(self):

        # print("Update Thread start OK!!!")

        sql_con = pymysql.connect(host=GlobalValues.SqlHostname,
                                  port=GlobalValues.SqlPort,
                                  user=GlobalValues.SqlUserName,
                                  passwd=GlobalValues.SqlPwd, db=GlobalValues.SqlDBName)

        while True:


            GlobalValues.checkThOTableUpdate = True


            try:
                if GlobalValues.checkOpenSpecFromGlobalStat == True:
                    GlobalValues.checkOpenSpecFromGlobalStat = False
                    self.btn_Spec.click()

                if (GlobalValues.checkGlobalviewMode == False):
                    if GlobalValues.TreeParentName == "Склад Офис":
                        self.btn_Spec.setEnabled(False)
                        self.btn_Documentation.setEnabled(False)
                        self.btn_PhotoArchive.setEnabled(False)
                        self.btn_TransportDocs.setEnabled(False)
                        self.btn_NetworkConfig.setEnabled(False)
                        self.btn_Statistic.setEnabled(False)
                    else:
                        self.btn_Spec.setEnabled(True)
                        self.btn_Documentation.setEnabled(True)
                        self.btn_PhotoArchive.setEnabled(True)
                        self.btn_TransportDocs.setEnabled(True)
                        self.btn_NetworkConfig.setEnabled(True)
                        self.btn_Statistic.setEnabled(True)
                else:
                    self.btn_Spec.setEnabled(False)
                    self.btn_Documentation.setEnabled(False)
                    self.btn_PhotoArchive.setEnabled(False)
                    self.btn_TransportDocs.setEnabled(False)
                    self.btn_NetworkConfig.setEnabled(False)
                    self.btn_Statistic.setEnabled(False)

                if GlobalValues.StopAll == True:
                    sql_con.close()
                    break
                else:
                    while GlobalValues.CHECK_TREE_UPDATE == True:
                        # print('waiting for tree updating is ending...')
                        time.sleep(0.05)

                    if ((GlobalValues.checkOTblUpdateEvent == True) and (GlobalValues.SqlHostname != '')) or (GlobalValues.SqlFirstConnectUpdate == True):
                        GlobalValues.checkOTblUpdateEvent = False

                        try:
                            # print("start oTbl Update!!!")

                            if (sql_con.open == False):
                                sql_con = pymysql.connect(host=GlobalValues.SqlHostname,
                                                          port=GlobalValues.SqlPort,
                                                          user=GlobalValues.SqlUserName,
                                                          passwd=GlobalValues.SqlPwd, db=GlobalValues.SqlDBName)

                            with sql_con:
                                cur = sql_con.cursor()
                                cur.execute("SELECT ID, OType, OModel, OSerialNum, OMaker, ODistributor, OComeTime, OTargetObj, OFactObj, OGoneToObjTime, OComment FROM ostuff")

                                rows = cur.fetchall()
                                time.sleep(0.5)

                                self.tbl_OStuff.setRowCount(0)

                                valWdg = 0

                                countRow = len(rows)

                                delay = round(2/countRow, 5)


                                for row in reversed(rows):
                                    # print('countrow :', str(countRow))
                                    # print('row', str(row))
                                    countRow -= 1
                                    self.tbl_OStuff.insertRow(valWdg)

                                    item = QtWidgets.QTableWidgetItem(str(row[0]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(valWdg, 0, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[1]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(valWdg, 1, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[2]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(valWdg, 2, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[3]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(valWdg, 3, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[4]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(valWdg, 4, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[5]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(valWdg, 5, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[6]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(valWdg, 6, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[7]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(valWdg, 7, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[8]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(valWdg, 8, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[9]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(valWdg, 9, item)

                                    # print('check')

                                    noneFlag = False
                                    if row[10] == None:
                                        item = QtWidgets.QTableWidgetItem("")
                                        noneFlag = True
                                    else:
                                        item = QtWidgets.QTableWidgetItem(row[10])
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    # print('valvdg', valWdg)
                                    self.tbl_OStuff.setItem(valWdg, 10, item)
                                    # print('valvdg', valWdg)
                                    if noneFlag == False:
                                        item.setToolTip(
                                            "<font color=black>%s</font>" % str(row[10]).replace("\n", "<br/>"))
                                    else:
                                        noneFlag = False

                                    valWdg += 1

                                    # time.sleep(delay)

                                cur.close()

                            # print('finish oTbl Update!')


                        except Exception:
                            print("UPDATE OTBL FAIL!!!", str(Exception))
                        time.sleep(0.1)
                        th_scroll = threading.Thread(target=self.thChangeScrollOTbl)
                        th_scroll.start()

                    if GlobalValues.checkTreeUpdateEvent == True:
                        # print('tree click!')
                        GlobalValues.checkTreeUpdateEvent = False

                        self.setOTblByTreeClick()
                        self.setKTblByTreeClick()

                        th_scroll = threading.Thread(target=self.thChangeScrollOTbl)
                        th_scroll.start()
                        th_scroll = threading.Thread(target=self.thChangeScrollKTbl)
                        th_scroll.start()


                    if ((GlobalValues.checkKTblUpdateEvent == True) and (GlobalValues.SqlHostname != '')) or (GlobalValues.SqlFirstConnectUpdate == True):
                        GlobalValues.checkKTblUpdateEvent = False
                        GlobalValues.SqlFirstConnectUpdate = False

                        try:
                            # print("REFRESH!!!")

                            if (sql_con.open == False):
                                sql_con = pymysql.connect(host=GlobalValues.SqlHostname,
                                                          port=GlobalValues.SqlPort,
                                                          user=GlobalValues.SqlUserName,
                                                          passwd=GlobalValues.SqlPwd, db=GlobalValues.SqlDBName)

                            with sql_con:
                                cur = sql_con.cursor()
                                cur.execute("SELECT ID, KType, KModel, KSerialNum, KMaker, KDistributor, KComeTime, KTargetObj, KFactObj, KGoneToObjTime, KComment FROM kstuff")

                                rows = cur.fetchall()
                                # print('AFTER SELECT')
                                pLenRows = len(rows)
                                time.sleep(0.1)
                                ppLenRows = len(rows)
                                # print(pLenRows)
                                # print(ppLenRows)

                                self.tbl_KStuff.setRowCount(0)

                                valWdg = 0

                                countRow = len(rows)

                                delay = round(2/countRow, 5)

                                for row in reversed(rows):
                                    # print('countrow :', str(countRow))
                                    # print('row', str(row))
                                    countRow -= 1
                                    self.tbl_KStuff.insertRow(valWdg)

                                    item = QtWidgets.QTableWidgetItem(str(row[0]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_KStuff.setItem(valWdg, 0, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[1]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_KStuff.setItem(valWdg, 1, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[2]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_KStuff.setItem(valWdg, 2, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[3]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_KStuff.setItem(valWdg, 3, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[4]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_KStuff.setItem(valWdg, 4, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[5]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_KStuff.setItem(valWdg, 5, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[6]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_KStuff.setItem(valWdg, 6, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[7]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_KStuff.setItem(valWdg, 7, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[8]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_KStuff.setItem(valWdg, 8, item)

                                    item = QtWidgets.QTableWidgetItem(str(row[9]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_KStuff.setItem(valWdg, 9, item)

                                    # print('check')

                                    noneFlag = False
                                    if row[10] == None:
                                        item = QtWidgets.QTableWidgetItem("")
                                        noneFlag = True
                                    else:
                                        item = QtWidgets.QTableWidgetItem(row[10])
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    # print('valvdg', valWdg)
                                    self.tbl_KStuff.setItem(valWdg, 10, item)
                                    # print('valvdg', valWdg)
                                    if noneFlag == False:
                                        item.setToolTip(
                                            "<font color=black>%s</font>" % str(row[10]).replace("\n", "<br/>"))
                                    else:
                                        noneFlag = False

                                    valWdg += 1

                                    # time.sleep(delay)

                                cur.close()

                            if GlobalValues.checkTreeStatesUpdate == True:
                                GlobalValues.checkTreeStatesUpdate = False
                                self.btn_RefreshTreeStates.click()

                        except Exception:
                            print("UPDATE OTBL FAIL!!!", str(Exception))
                            self.RefreshKTbl()
                        time.sleep(0.1)

                        th_scroll = threading.Thread(target=self.thChangeScrollKTbl)
                        th_scroll.start()

                    if GlobalValues.checkTreeStatesUpdate == True:
                        GlobalValues.checkTreeStatesUpdate = False
                        while GlobalValues.CHECK_OTBL_UPDATE or GlobalValues.CHECK_KTBL_UPDATE:
                            print('waiting')
                        print('print BTN States Refresh!')
                        self.btn_RefreshTreeStates.click()

                    if GlobalValues.checkTreeReBuild == True:
                        GlobalValues.checkTreeReBuild = False
                        self.treeUpdate()
                        time.sleep(0.2)
                        self.btn_RefreshTreeStates.click()

                    time.sleep(0.05)


            except Exception as ex:
                self.thUpdateOTable()
                print("ERROR THREAD: ", str(ex))

    def RefreshOTbl(self):
        try:
            print("REFRESH!!!")
            sql_con = pymysql.connect(host=GlobalValues.SqlHostname,
                                      port=GlobalValues.SqlPort,
                                      user=GlobalValues.SqlUserName,
                                      passwd=GlobalValues.SqlPwd, db=GlobalValues.SqlDBName)

            with sql_con:
                cur = sql_con.cursor()
                cur.execute("SELECT ID, OType, OModel, OSerialNum, OMaker, ODistributor, OComeTime, OTargetObj, OFactObj, OGoneToObjTime, OComment FROM ostuff")

                rows = cur.fetchall()
                # print('AFTER SELECT')
                time.sleep(0.2)

                self.tbl_OStuff.setRowCount(0)

                valWdg = 0

                countRow = len(rows)
                # print('countROWS', countRow)

                for row in reversed(rows):
                    # print('countrow :', str(countRow))
                    # print('row', str(row))
                    countRow -= 1
                    self.tbl_OStuff.insertRow(valWdg)


                    item = QtWidgets.QTableWidgetItem(str(row[0]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 0, item)

                    item = QtWidgets.QTableWidgetItem(str(row[1]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 1, item)

                    item = QtWidgets.QTableWidgetItem(str(row[2]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 2, item)

                    item = QtWidgets.QTableWidgetItem(str(row[3]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 3, item)

                    item = QtWidgets.QTableWidgetItem(str(row[4]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 4, item)

                    item = QtWidgets.QTableWidgetItem(str(row[5]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 5, item)

                    item = QtWidgets.QTableWidgetItem(str(row[6]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 6, item)

                    item = QtWidgets.QTableWidgetItem(str(row[7]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 7, item)

                    item = QtWidgets.QTableWidgetItem(str(row[8]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 8, item)

                    item = QtWidgets.QTableWidgetItem(str(row[9]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 9, item)

                    # print('check')

                    noneFlag = False
                    if row[10] == None:
                        item = QtWidgets.QTableWidgetItem("")
                        noneFlag = True
                    else:
                        item = QtWidgets.QTableWidgetItem(row[10])
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    # print('valvdg', valWdg)
                    self.tbl_OStuff.setItem(valWdg, 10, item)
                    # print('valvdg', valWdg)
                    if noneFlag == False:
                        item.setToolTip("<font color=black>%s</font>" % str(row[10]).replace("\n", "<br/>"))
                    else:
                        noneFlag = False

                    valWdg += 1
                # print('PRE CLOSE')
                cur.close()
            sql_con.close()
        except Exception:
            print("UPDATE OTBL FAIL!!!", str(Exception))
            self.RefreshOTbl()

    def sync_func(self):
        self.tbl_OStuff.verticalScrollBar().setValue(self.vScrl_OTbl.value())

    def thChangeScrollOTbl(self):
        try:

            num_delta =2000

            start_time = round(time.time() * 100)
            while True:
                numRows = self.tbl_OStuff.verticalScrollBar().maximum()
                delta = round(abs(time.time() * 100) - start_time)
                if numRows != 0:
                    time.sleep(0.2)
                    self.vScrl_OTbl.setMaximum(self.tbl_OStuff.verticalScrollBar().maximum())
                    break

                if delta > num_delta:
                    self.vScrl_OTbl.setMaximum(self.tbl_OStuff.verticalScrollBar().maximum())
                    break

                time.sleep(0.1)

        except:
            GlobalValues.checkVScrollOTblError = True

    def DeleteORecord(self):
        checkNull = -1
        if (int(self.tbl_OStuff.currentIndex().row()) == checkNull):
            GlobalValues.MesText = "Выберите оборудованиeе для удаления!"
            self.openMesBox()
        else:
            self.openPanelDelReason()
            if GlobalValues.checkDelComment == True:
                GlobalValues.checkDelComment = False
                if GlobalValues.MesResult == True:
                    GlobalValues.MesResult = False
                    con = pymysql.connect(host=str(GlobalValues.SqlHostname),
                                          port=int(GlobalValues.SqlPort),
                                          user=str(GlobalValues.SqlUserName),
                                          passwd=str(GlobalValues.SqlPwd),
                                          db=str(GlobalValues.SqlDBName))

                    OID = int(GlobalValues.OTblCurrID)
                    #print(str(OID))
                    with con:

                        cur = con.cursor()

                        query = ("DELETE FROM ostuff where id = (%s)")
                        cur.execute(query,(OID))
                        con.commit()
                        GlobalValues.Hystory(GlobalValues.UserName, GlobalValues.OTblCurrTypeObj,
                                             GlobalValues.OTblCurrSerialNumObj,
                                             "Удаление. Причина: " + str(GlobalValues.DelReasonComment))
                    if GlobalValues.checkGlobalviewMode == False:
                        GlobalValues.checkTreeUpdateEvent = True
                    else:
                        GlobalValues.checkOTblUpdateEvent = True
                    GlobalValues.checkTreeStatesUpdate = True


        # print('EDN')

    def cell_was_clicked(self):
        row = self.tbl_OStuff.currentIndex().row()
        col = self.tbl_OStuff.currentIndex().column()
        GlobalValues.OTblCurrTransferDateObj = ""
        itemID = self.tbl_OStuff.item(row, 0).text()
        typeObj = self.tbl_OStuff.item(row, 1).text()
        modelObj = self.tbl_OStuff.item(row, 2).text()
        serialNumObj = self.tbl_OStuff.item(row, 3).text()
        makerObj = self.tbl_OStuff.item(row, 4).text()
        distributorObj = self.tbl_OStuff.item(row, 5).text()
        inputDateObj = self.tbl_OStuff.item(row, 6).text()
        targetObj = self.tbl_OStuff.item(row, 7).text()
        factObj = self.tbl_OStuff.item(row, 8).text()
        transferDateObj = self.tbl_OStuff.item(row, 9).text()
        comment = self.tbl_OStuff.item(row,10).text()
        self.OTblItemID = itemID
        GlobalValues.OTblCurrID = self.OTblItemID
        GlobalValues.OTblCurrTypeObj = str(typeObj)
        GlobalValues.OTblCurrModelObj = str(modelObj)
        GlobalValues.OTblCurrSerialNumObj = str(serialNumObj)
        GlobalValues.OTblCurrMakerObj = str(makerObj)
        GlobalValues.OTblCurrDistributorObj = str(distributorObj)
        GlobalValues.OTblCurrInputDateObj = str(inputDateObj)
        GlobalValues.OTblCurrTargetObj = str(targetObj)
        GlobalValues.OTblCurrFactObj = str(factObj)
        GlobalValues.OTblCurrTransferDateObj = str(transferDateObj)
        GlobalValues.OTblCurrComment = str(comment)
        GlobalValues.OTblCurrRow = str(row)
        GlobalValues.OTblCurrCol = str(col)

        GlobalValues.checkOTblWasClicked = True

    def AddNewObject(self):
        TreeItemCount = self.tree_Obj.topLevelItemCount()
        item_0 = QtWidgets.QTreeWidgetItem(self.tree_Obj)
        item_0.setToolTip(0, GlobalValues.NewObjName)
        self.tree_Obj.topLevelItem(TreeItemCount).setText(0,str(GlobalValues.NewObjName))

        locHostName = str(GlobalValues.SqlHostname)
        locPort = int(GlobalValues.SqlPort)
        locUserName = str(GlobalValues.SqlUserName)
        locPwd = str(GlobalValues.SqlPwd)
        locDBName = str(GlobalValues.SqlDBName)
        con = pymysql.connect(host=locHostName,
                              port=locPort,
                              user=locUserName,
                              passwd=locPwd,
                              db=locDBName)
        with con:

            cur = con.cursor()
            cur.execute("SELECT ObjID FROM treeobjtbl")
            rows = cur.fetchall()
            #
            ObjID = 1
            #print(len(rows))
            if (len(rows) != 0):
                for row in reversed(rows):
                    try:
                        ObjID = str(int(row[0]) + 1)
                    except:
                        print("OID Error!!!")
                    break
            ParentName = GlobalValues.NewObjName
            ADD_query = ("INSERT INTO treeobjtbl (ObjID, ParentObjName, ConnectionID) VALUES ( %s, %s, %s)")
            cur.execute(ADD_query, (ObjID, ParentName, ObjID))
            con.commit()

        con.close()
        self.btn_RefreshTreeStates.click()

    def DelTreeObj(self):
        checkParentStuff = False
        checkChildStuff = False
        checkDelAccept = False
        ostuffcount = 0
        kstuffcount = 0
        curr_conID = 0
        curr_PconID = 0
        GlobalValues.MesText = "Удалить объект. Вы уверены?"
        self.openMesBox()
        if (GlobalValues.MesResult == True):
            root = self.tree_Obj.invisibleRootItem()
            for item in self.tree_Obj.selectedItems():
                # print(item.text(0))
                if (str(item.text(0)) == "Склад Офис"):
                    GlobalValues.MesText = "Объект нельзя удалить!!!"
                    self.openMesBox()
                else:

                    ObjName = item.text(0)
                    # print((ObjName))
                    locHostName = str(GlobalValues.SqlHostname)
                    locPort = int(GlobalValues.SqlPort)
                    locUserName = str(GlobalValues.SqlUserName)
                    locPwd = str(GlobalValues.SqlPwd)
                    locDBName = str(GlobalValues.SqlDBName)
                    con = pymysql.connect(host=locHostName,
                                          port=locPort,
                                          user=locUserName,
                                          passwd=locPwd,
                                          db=locDBName)
                    with con:

                        D_cur = con.cursor()
                        D_cur.execute("SELECT ParentObjName, ConnectionID FROM treeobjtbl")
                        ParentName_rows = D_cur.fetchall()
                        D_cur.execute("SELECT ChildID, ChildNameObj, ConnectionID FROM treechildobjtbl")
                        ChildName_rows = D_cur.fetchall()

                        curr_conID = 0
                        #print(ParentName_rows)
                        for row in ParentName_rows:
                            if str(row[0]) == str(ObjName):
                                curr_conID = row[1]
                                checkParentStuff = True
                        if curr_conID == 0:
                            for row2 in ChildName_rows:
                                if str(row2[1]) == str(ObjName):
                                    curr_conID = row2[2]
                                    curr_PconID = row2[0]
                                    checkChildStuff = True
                        # print("PID: ", str(curr_PconID))
                        # print("CID: ", str(curr_conID))
                        # print("childDelCheck: ", str(checkChildStuff))

                        if checkParentStuff == True:
                            D_cur.execute("SELECT ParentID FROM ostuff WHERE ParentID Like '" + str(curr_conID) + "'")
                            ParentOStuff_rows = D_cur.fetchall()
                            D_cur.execute("SELECT ParentID FROM kstuff WHERE ParentID Like '" + str(curr_conID) + "'")
                            ParentKStuff_rows = D_cur.fetchall()
                            if (len(ParentOStuff_rows) > 0) or (len(ParentKStuff_rows) > 0):
                                # print("ID = ", str(curr_conID))
                                # print("Parent ERR : " + str(len(ParentOStuff_rows)) + "|" + str(len(ParentKStuff_rows)))
                                GlobalValues.MesText = "Объект нельзя удалить!"
                                self.openMesBox()
                                checkDelAccept = True

                        if checkChildStuff == True:
                            D_cur.execute("SELECT ParentID, ChildID FROM ostuff WHERE ParentID Like '" + str(curr_conID) + "'")
                            ChildOStuff_rows = D_cur.fetchall()
                            D_cur.execute("SELECT ParentID, ChildID FROM kstuff WHERE ParentID Like '" + str(curr_conID) + "'")
                            ChildKStuff_rows = D_cur.fetchall()

                            if (len(ChildOStuff_rows) > 0):
                                for orow in ChildOStuff_rows:
                                    if str(orow[1]) == str(curr_PconID):
                                        ostuffcount+=1
                            if (len(ChildKStuff_rows) > 0):

                                for krow in ChildKStuff_rows:
                                    if str(krow[1] == str(curr_PconID)):
                                        kstuffcount+=1
                            # print("lenRows: ", str(len(ChildOStuff_rows)), " | ", str(len(ChildKStuff_rows)))
                            # print("counters: ", str(ostuffcount), " | ", str(kstuffcount) )


                            if (ostuffcount != 0) or (kstuffcount != 0):
                                # print("Child ERR : " + str(len(ChildOStuff_rows)) + "|" + str(len(ChildKStuff_rows)))
                                GlobalValues.MesText = "Объект нельзя удалить!"
                                self.openMesBox()
                                checkDelAccept = True

                        if checkDelAccept == False:
                            (item.parent() or root).removeChild(item)
                            Parent_query = ("DELETE FROM treeobjtbl where ParentObjName = (%s)")
                            D_cur.execute(Parent_query, (str(ObjName)))

                            Child_query = ("DELETE FROM treechildobjtbl where ConnectionID = (%s)")
                            D_cur.execute(Child_query, (str(curr_conID)))
            self.btn_RefreshTreeStates.click()

    def AddNewChildObject(self):


        text, ok = QInputDialog.getText(self, "Добавление подобъекта", "Имя подобъекта:")
        if ok and text != "":
            if len(self.tree_Obj.selectedItems()) > 0:
                QTreeWidgetItem(self.tree_Obj.selectedItems()[0], [text])

                locHostName = str(GlobalValues.SqlHostname)
                locPort = int(GlobalValues.SqlPort)
                locUserName = str(GlobalValues.SqlUserName)
                locPwd = str(GlobalValues.SqlPwd)
                locDBName = str(GlobalValues.SqlDBName)
                con = pymysql.connect(host=locHostName,
                                      port=locPort,
                                      user=locUserName,
                                      passwd=locPwd,
                                      db=locDBName)


                curr_selected_item = self.tree_Obj.currentItem()


                with con:

                    cur = con.cursor()
                    cur.execute("SELECT ChildID FROM treechildobjtbl")
                    child_rows = cur.fetchall()
                    ObjChildID = 1
                    if (len(child_rows) != 0):
                        for jrow in reversed(child_rows):
                            try:
                                ObjChildID = str(int(jrow[0]) + 1)
                            except:
                                print("OID Error!!!")
                            break

                    ChildName = text

                    Parent_cur_conID = con.cursor()
                    Parent_cur_conID.execute("SELECT ParentObjName, ConnectionID FROM treeobjtbl")
                    Parent_conID_rows = Parent_cur_conID.fetchall()
                    ChildConID = 0
                    curr_selected_item = self.tree_Obj.currentItem()
                    parent_item = str(curr_selected_item.text(0))
                    #print(str(curr_selected_item.text(0)))
                    for irow in Parent_conID_rows:
                        if str(irow[0]) == parent_item:
                            ChildConID = irow[1]


                    ParentName = GlobalValues.NewObjName
                    ADD_child_query = ("INSERT INTO treechildobjtbl (ChildID, ChildNameObj, ConnectionID) VALUES ( %s, %s, %s)")
                    cur.execute(ADD_child_query, (ObjChildID, ChildName, ChildConID))
                    con.commit()

                con.close()
            else:
                QTreeWidgetItem(self.tree_Obj, [text])

            self.btn_RefreshTreeStates.click()

    def openMesBox(self):
        self.uiMesBox = Ui_MesBox()
        self.uiMesBox.exec_()

    def openPanelAddNewObj(self):
        GlobalValues.DialogHeadText = "Имя объекта: "
        GlobalValues.DialogText = ""
        self.openDialogBox()
        if GlobalValues.DialogRes == True:
            GlobalValues.NewObjName = GlobalValues.DialogText
            self.AddNewObject()

    def treeUpdate(self):
        print('Start tree Update')
        try:
            GlobalValues.CHECK_TREE_UPDATE = True
            print('Glob Flag: ', str(GlobalValues.checkGlobalviewMode))
            if GlobalValues.checkGlobalviewMode == False:
                temp_currTreeItemName = GlobalValues.TreeParentName
            self.tree_Obj.clear()
            locHostName = str(GlobalValues.SqlHostname)
            locPort = int(GlobalValues.SqlPort)
            locUserName = str(GlobalValues.SqlUserName)
            locPwd = str(GlobalValues.SqlPwd)
            locDBName = str(GlobalValues.SqlDBName)
            con = pymysql.connect(host=locHostName,
                                  port=locPort,
                                  user=locUserName,
                                  passwd=locPwd,
                                  db=locDBName)

            # curr_selected_item = self.tree_Obj.currentItem()

            with con:
                temp_Parent_cur = con.cursor()
                temp_Parent_cur.execute("SELECT ParentObjName, ConnectionID FROM treeobjtbl")
                temp_Parent_rows = temp_Parent_cur.fetchall()
                #print("123")
                index = 1
                temp_Child_cur = con.cursor()
                temp_Child_cur.execute("SELECT ChildNameObj, ConnectionID FROM treechildobjtbl")
                temp_Child_rows = temp_Child_cur.fetchall()
                newtext = "111"
                _translate = QtCore.QCoreApplication.translate
                parent_index = 0
                i = 0
                for row in temp_Parent_rows:
                    if str(row[0]) != "Резерв":
                        TreeItemCount = self.tree_Obj.topLevelItemCount()
                        item_0 = QtWidgets.QTreeWidgetItem(self.tree_Obj)
                        item_0.setToolTip(0, str(row[0]))

                        GlobalValues.TreeParentName = str(row[0])
                        self.tree_Obj.topLevelItem(TreeItemCount).setText(0, str(row[0]))

                        #back-styles
                        # if str(row[0]) != "Склад Офис":
                        #     checkObj = self.getTreeObjStat()
                        #     if checkObj == '2':
                        #         self.tree_Obj.topLevelItem(TreeItemCount).setBackground(0, QtGui.QBrush(self.green_gradient))
                        #     else:
                        #         if checkObj == '1':
                        #             self.tree_Obj.topLevelItem(TreeItemCount).setBackground(0, QtGui.QBrush(self.yellow_gradient))
                        #         else:
                        #             self.tree_Obj.topLevelItem(TreeItemCount).setBackground(0, QtGui.QBrush(self.red_gradient))


                        index += 1
                        child_index = 0
                        for child_row in temp_Child_rows:
                            if str(row[1]) == str(child_row[1]):
                                childtext = child_row[0]
                                #print(childtext)
                                item_1 = QtWidgets.QTreeWidgetItem(item_0)
                                self.tree_Obj.topLevelItem(parent_index).child(child_index).setText(0,str(child_row[0]))
                                child_index+=1
                        parent_index+=1
                        i+=1

                if GlobalValues.checkGlobalviewMode == False:
                    count = self.tree_count_tems()
                    i = 0
                    for i in range(count):
                        self.tree_Obj.setCurrentItem(self.tree_Obj.topLevelItem(i))
                        item = self.tree_Obj.currentItem()
                        itemText = item.text(0)
                        if itemText == temp_currTreeItemName:
                            break
                        i += 1
                    print('BTN States Refresh by KTBL end!')
            print('End tree Update')
            GlobalValues.CHECK_TREE_UPDATE = False
        except:
            GlobalValues.checkFirstConn = False

    def setOTblByTreeClick(self):
        GlobalValues.CHECK_OTBL_UPDATE = True
        # print("O TBL BY TREE REFRESH!!!")
        # self.tbl_OStuff.clear()
        self.tbl_OStuff.setRowCount(0)
        item = self.tree_Obj.currentItem()
        currItemText = item.text(0)
        # print(currItemText)
        checkParentTransfer = False
        # print(str(currItemText))

        con = pymysql.connect(host=str(GlobalValues.SqlHostname),
                              port=int(GlobalValues.SqlPort),
                              user=str(GlobalValues.SqlUserName),
                              passwd=str(GlobalValues.SqlPwd),
                              db=str(GlobalValues.SqlDBName))
        with con:
            cur = con.cursor()

            cur.execute("SELECT ObjID, ParentObjName, ConnectionID FROM treeobjtbl")
            rows = cur.fetchall()
            cur.execute("SELECT ID, OType, OModel, OSerialNum, OMaker, ODistributor, OComeTime, OTargetObj, OFactObj, OGoneToObjTime, ParentID, ChildID, OComment FROM ostuff")
            Parent_rows = cur.fetchall()
            time.sleep(0.2)

            for row in rows:
                if str(row[1]) == str(currItemText):
                    checkParentTransfer = True
                    currParentID = str(row[0])
                    countRow = 0
                    for temp1_row in Parent_rows:
                        countRow += 1
                    valWdg = 0
                    # if str(temp1_row[9]) == str(currItemText):

                    for temp2_row in reversed(Parent_rows):
                        #     countRow -= 1
                        if str(temp2_row[8]) == str(currItemText):
                            # countRow -= 1
                            self.tbl_OStuff.insertRow(valWdg)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[0]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_OStuff.setItem(valWdg, 0, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[1]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_OStuff.setItem(valWdg, 1, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[2]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_OStuff.setItem(valWdg, 2, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[3]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_OStuff.setItem(valWdg, 3, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[4]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_OStuff.setItem(valWdg, 4, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[5]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_OStuff.setItem(valWdg, 5, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[6]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_OStuff.setItem(valWdg, 6, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[7]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_OStuff.setItem(valWdg, 7, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[8]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_OStuff.setItem(valWdg, 8, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[9]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_OStuff.setItem(valWdg, 9, item)

                            noneFlag = False
                            if temp2_row[12] == None:
                                item = QtWidgets.QTableWidgetItem("")
                                noneFlag = True
                            else:
                                item = QtWidgets.QTableWidgetItem(temp2_row[12])
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_OStuff.setItem(valWdg, 10, item)
                            if noneFlag == False:
                                item.setToolTip("<font color=black>%s</font>" % str(temp2_row[12]).replace("\n", "<br/>"))
                            else:
                                noneFlag = False


                            valWdg += 1
            if checkParentTransfer == False:
                cur.execute("SELECT ChildID, ChildNameObj, ConnectionID FROM treechildobjtbl")
                childList_rows = cur.fetchall()

                # for row1 in childList_rows:
                #     if
                #
                # curChildID = ''
                # for row1 in childList_rows:
                #     if str(row1[1]) == str(currItemText):
                #         curChildID = str(row1[0])
                #         print('CHID == ' + curChildID)
                for temp1_child_row in childList_rows:
                        ChildvalWdg = 0
                        if str(temp1_child_row[1]) == str(currItemText):
                            currChildID = str(temp1_child_row[0])
                            # print(str(currChildID))
                            for temp2_child_row in Parent_rows:
                                if (str(temp2_child_row[11]) == str(currChildID) and str(temp2_child_row[8]) == str(GlobalValues.TreeParentName)):
                                    # print([str(temp2_child_row[11])])
                                    self.tbl_OStuff.insertRow(ChildvalWdg)

                                    item = QtWidgets.QTableWidgetItem(str(temp2_child_row[0]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(ChildvalWdg, 0, item)

                                    item = QtWidgets.QTableWidgetItem(str(temp2_child_row[1]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(ChildvalWdg, 1, item)

                                    item = QtWidgets.QTableWidgetItem(str(temp2_child_row[2]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(ChildvalWdg, 2, item)

                                    item = QtWidgets.QTableWidgetItem(str(temp2_child_row[3]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(ChildvalWdg, 3, item)

                                    item = QtWidgets.QTableWidgetItem(str(temp2_child_row[4]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(ChildvalWdg, 4, item)

                                    item = QtWidgets.QTableWidgetItem(str(temp2_child_row[5]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(ChildvalWdg, 5, item)

                                    item = QtWidgets.QTableWidgetItem(str(temp2_child_row[6]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(ChildvalWdg, 6, item)

                                    item = QtWidgets.QTableWidgetItem(str(temp2_child_row[7]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(ChildvalWdg, 7, item)

                                    item = QtWidgets.QTableWidgetItem(str(temp2_child_row[8]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(ChildvalWdg, 8, item)

                                    item = QtWidgets.QTableWidgetItem(str(temp2_child_row[9]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(ChildvalWdg, 9, item)

                                    noneFlag = False
                                    if temp2_child_row[12] == None:
                                        item = QtWidgets.QTableWidgetItem("")
                                        noneFlag = True
                                    else:
                                        item = QtWidgets.QTableWidgetItem(temp2_child_row[12])
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tbl_OStuff.setItem(ChildvalWdg, 10, item)
                                    if noneFlag == False:
                                        item.setToolTip(
                                            "<font color=black>%s</font>" % str(temp2_child_row[12]).replace("\n", "<br/>"))
                                    else:
                                        noneFlag = False

                                    ChildvalWdg += 1

        th_scroll = threading.Thread(target=self.thChangeScrollOTbl)
        th_scroll.start()

        GlobalValues.CHECK_OTBL_UPDATE = False

    def BackToMain(self):
        #сброс таблицы otbl
        self.le_OViewMode.setText("Глобальный")
        self.tree_Obj.clearSelection()
        GlobalValues.checkGlobalviewMode = True
        GlobalValues.checkOTblUpdateEvent = True
        #сброс таблицы ktbl
        self.le_KViewMode.setText("Глобальный")
        self.tree_Obj.clearSelection()
        GlobalValues.checkGlobalviewMode = True
        GlobalValues.checkKTblUpdateEvent = True

    def FirstOTblUpdate(self):
        try:
            sql_con = pymysql.connect(host=GlobalValues.SqlHostname,
                                      port=GlobalValues.SqlPort,
                                      user=GlobalValues.SqlUserName,
                                      passwd=GlobalValues.SqlPwd, db=GlobalValues.SqlDBName)

            with sql_con:
                cur = sql_con.cursor()
                cur.execute(
                    "SELECT ID, OType, OModel, OSerialNum, OMaker, ODistributor, OComeTime, OTargetObj, OFactObj, OGoneToObjTime, OComment FROM ostuff")

                rows = cur.fetchall()
                time.sleep(0.2)
                # print(str(row))

                self.tbl_OStuff.setRowCount(0)

                valWdg = 0

                countRow = 0
                for row in rows:
                    countRow += 1

                for row in reversed(rows):
                    countRow -= 1
                    self.tbl_OStuff.insertRow(valWdg)

                    item = QtWidgets.QTableWidgetItem(str(row[0]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 0, item)

                    item = QtWidgets.QTableWidgetItem(str(row[1]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 1, item)

                    item = QtWidgets.QTableWidgetItem(str(row[2]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 2, item)

                    item = QtWidgets.QTableWidgetItem(str(row[3]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 3, item)

                    item = QtWidgets.QTableWidgetItem(str(row[4]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 4, item)

                    item = QtWidgets.QTableWidgetItem(str(row[5]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 5, item)

                    item = QtWidgets.QTableWidgetItem(str(row[6]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 6, item)

                    item = QtWidgets.QTableWidgetItem(str(row[7]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 7, item)

                    item = QtWidgets.QTableWidgetItem(str(row[8]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 8, item)

                    item = QtWidgets.QTableWidgetItem(str(row[9]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 9, item)

                    noneFlag = False
                    if row[10] == None:
                        item = QtWidgets.QTableWidgetItem("")
                        noneFlag = True
                    else:
                        item = QtWidgets.QTableWidgetItem(row[10])
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 10, item)
                    if noneFlag == False:
                        item.setToolTip("<font color=black>%s</font>" % str(row[10]).replace("\n", "<br/>"))
                    else:
                        noneFlag = False
        except:
            GlobalValues.checkFirstConn = False
            GlobalValues.MesText = "Подключение к БД отсутствует!"
            self.openMesBox()

    def openPanelEditOStuff(self):
        if GlobalValues.checkOTblWasClicked == True :
            if GlobalValues.checkGroupOTbl == True:

                temp_lst_selected = []
                lst_selected = []
                selected = self.tbl_OStuff.selectedItems()
                for item in selected:

                    if item.column() == 0:
                        temp_lst_selected.append(item.data(0))
                    if item.column() == 1:
                        temp_lst_selected.append(item.data(0))
                    if item.column() == 3:
                        temp_lst_selected.append(item.data(0))
                    if item.column() == 10:
                        temp_lst_selected.append(item.data(0))
                        lst_selected.append(temp_lst_selected)
                        temp_lst_selected = []
                GlobalValues.otbl_selected_lst = lst_selected
                self.uiOGroupEdit = Ui_OGroupEdit()
                self.uiOGroupEdit.exec_()
            else:
                self.uiEditOStuff = Ui_EditOStuff()
                self.uiEditOStuff.exec_()
        else:
            GlobalValues.MesText = "Оборудование не выбрано!"
            self.openMesBox()

    def SearchEnable(self):
        if self.cb_OStuff.currentText() == "Нет" :
            self.le_OStuff.setEnabled(False)
        else:
            self.le_OStuff.setEnabled(True)

    def SearchOStuff(self):

        if (self.le_OStuff.text() == "") :
            GlobalValues.MesText = "Введите текст для поика!"
            self.openMesBox()
        else:
            self.tbl_OStuff.setRowCount(0)
            curSearchType = self.cb_OStuff.currentText()
            con = pymysql.connect(host=str(GlobalValues.SqlHostname),
                                  port=int(GlobalValues.SqlPort),
                                  user=str(GlobalValues.SqlUserName),
                                  passwd=str(GlobalValues.SqlPwd),
                                  db=str(GlobalValues.SqlDBName))
            curSearchText = '%' + str(self.le_OStuff.text()) + '%'
            # print(curSearchType)
            with con:
                cur = con.cursor()
                if curSearchType == "Серийный номер":
                    cur.execute(
                        "SELECT ID, OType, OModel, OSerialNum, OMaker, ODistributor, OComeTime, OTargetObj, OFactObj, OGoneToObjTime, ParentID, ChildID, OComment FROM ostuff WHERE OSerialNum Like '" + curSearchText + "'")
                if curSearchType == "Поставщик":
                    cur.execute(
                        "SELECT ID, OType, OModel, OSerialNum, OMaker, ODistributor, OComeTime, OTargetObj, OFactObj, OGoneToObjTime, ParentID, ChildID, OComment FROM ostuff WHERE ODistributor Like '" + curSearchText + "'")
                if curSearchType == "Цел.Объект":
                    cur.execute(
                        "SELECT ID, OType, OModel, OSerialNum, OMaker, ODistributor, OComeTime, OTargetObj, OFactObj, OGoneToObjTime, ParentID, ChildID, OComment FROM ostuff WHERE OTargetObj Like '" + curSearchText + "'")

                rows = cur.fetchall()
                valWdg = 0
                self.tbl_OStuff.setRowCount(0)
                for row in rows:

                    self.tbl_OStuff.insertRow(valWdg)

                    item = QtWidgets.QTableWidgetItem(str(row[0]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 0, item)

                    item = QtWidgets.QTableWidgetItem(str(row[1]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 1, item)

                    item = QtWidgets.QTableWidgetItem(str(row[2]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 2, item)

                    item = QtWidgets.QTableWidgetItem(str(row[3]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 3, item)

                    item = QtWidgets.QTableWidgetItem(str(row[4]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 4, item)

                    item = QtWidgets.QTableWidgetItem(str(row[5]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 5, item)

                    item = QtWidgets.QTableWidgetItem(str(row[6]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 6, item)

                    item = QtWidgets.QTableWidgetItem(str(row[7]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 7, item)

                    item = QtWidgets.QTableWidgetItem(str(row[8]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 8, item)

                    item = QtWidgets.QTableWidgetItem(str(row[9]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 9, item)

                    noneFlag = False
                    if row[12] == None:
                        item = QtWidgets.QTableWidgetItem("")
                        noneFlag = True
                    else:
                        item = QtWidgets.QTableWidgetItem(row[12])
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_OStuff.setItem(valWdg, 10, item)
                    if noneFlag == False:
                        item.setToolTip("<font color=black>%s</font>" % str(row[12]).replace("\n", "<br/>"))
                    else:
                        noneFlag = False

                    valWdg += 1
                cur.close()
            con.close()

    def OTblMode(self):
        if self.chB_OGroup.isChecked() == True:
            GlobalValues.checkGroupOTbl = True
            self.tbl_OStuff.clearSelection()
            self.tbl_OStuff.setSelectionMode(QtWidgets.QAbstractItemView.MultiSelection)
        else:
            GlobalValues.checkGroupOTbl = False
            self.tbl_OStuff.clearSelection()
            self.tbl_OStuff.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)

    def getStatOTbl(self):
        GlobalValues.CurrStatArray = self.getTblStat(self.tbl_OStuff)
        self.openPanelStat()
    #работа с комплектующими
    def openPanelAddNewKStuff(self):
        self.uiAddNewKStuff = Ui_AddNewKStuff()
        self.uiAddNewKStuff.exec_()

    def FirstKTblUpdate(self):
        # print("FIRST KTBL UPDATE!!!")
        try:
            sql_con = pymysql.connect(host=GlobalValues.SqlHostname,
                                      port=GlobalValues.SqlPort,
                                      user=GlobalValues.SqlUserName,
                                      passwd=GlobalValues.SqlPwd, db=GlobalValues.SqlDBName)

            with sql_con:
                cur = sql_con.cursor()
                cur.execute(
                    "SELECT ID, KType, KModel, KSerialNum, KMaker, KDistributor, KComeTime, KTargetObj, KFactObj, KGoneToObjTime, KComment FROM kstuff")

                rows = cur.fetchall()
                time.sleep(0.2)
                # print(str(row))

                self.tbl_KStuff.setRowCount(0)

                valWdg = 0

                countRow = 0
                for row in rows:
                    countRow += 1

                for row in reversed(rows):
                    countRow -= 1
                    self.tbl_KStuff.insertRow(valWdg)

                    item = QtWidgets.QTableWidgetItem(str(row[0]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 0, item)

                    item = QtWidgets.QTableWidgetItem(str(row[1]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 1, item)

                    item = QtWidgets.QTableWidgetItem(str(row[2]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 2, item)

                    item = QtWidgets.QTableWidgetItem(str(row[3]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 3, item)

                    item = QtWidgets.QTableWidgetItem(str(row[4]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 4, item)

                    item = QtWidgets.QTableWidgetItem(str(row[5]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 5, item)

                    item = QtWidgets.QTableWidgetItem(str(row[6]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 6, item)

                    item = QtWidgets.QTableWidgetItem(str(row[7]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 7, item)

                    item = QtWidgets.QTableWidgetItem(str(row[8]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 8, item)

                    item = QtWidgets.QTableWidgetItem(str(row[9]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 9, item)

                    noneFlag = False
                    if row[10] == None:
                        item = QtWidgets.QTableWidgetItem("")
                        noneFlag = True
                    else:
                        item = QtWidgets.QTableWidgetItem(row[10])
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 10, item)
                    if noneFlag == False:
                        item.setToolTip("<font color=black>%s</font>" % str(row[10]).replace("\n", "<br/>"))
                    else:
                        noneFlag = False

                    valWdg += 1
        except:
            GlobalValues.checkFirstConn = False

    def thUpdateKTable(self):
        # print("Update Func start OK!!!")
        th =threading.Thread(target = self.KTableUpdate)
        th.start()

    def KTableUpdate(self):

        # print("Update Thread start OK!!!")

        while True:
            GlobalValues.checkThKTableUpdate = True

            try:
                if (GlobalValues.checkKTblUpdateEvent == True) and (GlobalValues.SqlHostname != '') or (GlobalValues.SqlFirstConnectUpdate == True):
                    GlobalValues.checkKTblUpdateEvent = False
                    GlobalValues.SqlFirstConnectUpdate = False
                    self.RefreshKTbl()
                    th_scroll = threading.Thread(target=self.thChangeScrollKTbl)
                    th_scroll.start()
                if GlobalValues.checkTreeUpdateEvent == True:
                    GlobalValues.checkTreeUpdateEvent = False
                    self.setKTblByTreeClick()
                    th_scroll = threading.Thread(target=self.thChangeScrollKTbl)
                    th_scroll.start()

            except:
                GlobalValues.checkThKTableUpdate = False
                # print("123123123")

    def RefreshKTbl(self):
        #print("REFRESH!!!")
        sql_con = pymysql.connect(host=GlobalValues.SqlHostname,
                                  port=GlobalValues.SqlPort,
                                  user=GlobalValues.SqlUserName,
                                  passwd=GlobalValues.SqlPwd, db=GlobalValues.SqlDBName)

        with sql_con:
            cur = sql_con.cursor()
            cur.execute("SELECT ID, KType, KModel, KSerialNum, KMaker, KDistributor, KComeTime, KTargetObj, KFactObj, KGoneToObjTime, KComment FROM kstuff")

            rows = cur.fetchall()
            time.sleep(0.2)
            # print(str(row))

            self.tbl_KStuff.setRowCount(0)

            valWdg = 0

            countRow = 0
            for row in rows:
                countRow += 1

            for row in reversed(rows):
                countRow -= 1
                self.tbl_KStuff.insertRow(valWdg)


                item = QtWidgets.QTableWidgetItem(str(row[0]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                self.tbl_KStuff.setItem(valWdg, 0, item)

                item = QtWidgets.QTableWidgetItem(str(row[1]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                self.tbl_KStuff.setItem(valWdg, 1, item)

                item = QtWidgets.QTableWidgetItem(str(row[2]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                self.tbl_KStuff.setItem(valWdg, 2, item)

                item = QtWidgets.QTableWidgetItem(str(row[3]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                self.tbl_KStuff.setItem(valWdg, 3, item)

                item = QtWidgets.QTableWidgetItem(str(row[4]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                self.tbl_KStuff.setItem(valWdg, 4, item)

                item = QtWidgets.QTableWidgetItem(str(row[5]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                self.tbl_KStuff.setItem(valWdg, 5, item)

                item = QtWidgets.QTableWidgetItem(str(row[6]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                self.tbl_KStuff.setItem(valWdg, 6, item)

                item = QtWidgets.QTableWidgetItem(str(row[7]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                self.tbl_KStuff.setItem(valWdg, 7, item)

                item = QtWidgets.QTableWidgetItem(str(row[8]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                self.tbl_KStuff.setItem(valWdg, 8, item)

                item = QtWidgets.QTableWidgetItem(str(row[9]))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                self.tbl_KStuff.setItem(valWdg, 9, item)

                noneFlag = False
                if row[10] == None:
                    item = QtWidgets.QTableWidgetItem("")
                    noneFlag = True
                else:
                    item = QtWidgets.QTableWidgetItem(row[10])
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                self.tbl_KStuff.setItem(valWdg, 10, item)
                if noneFlag == False:
                    item.setToolTip("<font color=black>%s</font>" % str(row[10]).replace("\n", "<br/>"))
                else:
                    noneFlag = False

                valWdg += 1

    def sync_func2(self):
        self.tbl_KStuff.verticalScrollBar().setValue(self.vScrl_KTbl_2.value())

    def thChangeScrollKTbl(self):
        try:
            num_delta =2000

            start_time = round(time.time() * 100)
            while True:
                numRows = self.tbl_KStuff.verticalScrollBar().maximum()
                delta = round(abs(time.time() * 100) - start_time)
                # print(numRows)

                if numRows != 0:
                    time.sleep(0.2)
                    self.vScrl_KTbl_2.setMaximum(self.tbl_KStuff.verticalScrollBar().maximum())
                    break

                if delta > num_delta:
                    self.vScrl_KTbl_2.setMaximum(self.tbl_KStuff.verticalScrollBar().maximum())
                    break


                time.sleep(0.1)

        except:
            GlobalValues.checkVScrollKTblError = True

    def setKTblByTreeClick(self):
        GlobalValues.CHECK_KTBL_UPDATE = True
        # print("K TBL BY TREE REFRESH!!!")
        # self.tbl_OStuff.clear()
        self.tbl_KStuff.setRowCount(0)
        item = self.tree_Obj.currentItem()
        currItemText = item.text(0)
        checkParentTransfer = False
        # print(str(currItemText))

        con = pymysql.connect(host=str(GlobalValues.SqlHostname),
                              port=int(GlobalValues.SqlPort),
                              user=str(GlobalValues.SqlUserName),
                              passwd=str(GlobalValues.SqlPwd),
                              db=str(GlobalValues.SqlDBName))
        with con:
            cur = con.cursor()

            cur.execute("SELECT ObjID, ParentObjName, ConnectionID FROM treeobjtbl")
            rows = cur.fetchall()
            cur.execute("SELECT ID, KType, KModel, KSerialNum, KMaker, KDistributor, KComeTime, KTargetObj, KFactObj, KGoneToObjTime, ParentID, ChildID, KComment FROM kstuff")
            Parent_rows = cur.fetchall()
            time.sleep(0.2)

            for row in rows:
                if str(row[1]) == str(currItemText):
                    checkParentTransfer = True
                    currParentID = str(row[0])
                    countRow = 0
                    for temp1_row in Parent_rows:
                        countRow += 1
                    valWdg = 0
                    # if str(temp1_row[9]) == str(currItemText):

                    for temp2_row in reversed(Parent_rows):
                        #     countRow -= 1
                        if str(temp2_row[8]) == str(currItemText):
                            # countRow -= 1
                            self.tbl_KStuff.insertRow(valWdg)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[0]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_KStuff.setItem(valWdg, 0, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[1]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_KStuff.setItem(valWdg, 1, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[2]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_KStuff.setItem(valWdg, 2, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[3]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_KStuff.setItem(valWdg, 3, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[4]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_KStuff.setItem(valWdg, 4, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[5]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_KStuff.setItem(valWdg, 5, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[6]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_KStuff.setItem(valWdg, 6, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[7]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_KStuff.setItem(valWdg, 7, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[8]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_KStuff.setItem(valWdg, 8, item)

                            item = QtWidgets.QTableWidgetItem(str(temp2_row[9]))
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_KStuff.setItem(valWdg, 9, item)

                            noneFlag = False
                            if temp2_row[12] == None:
                                item = QtWidgets.QTableWidgetItem("")
                                noneFlag = True
                            else:
                                item = QtWidgets.QTableWidgetItem(temp2_row[12])
                            item.setTextAlignment(Qt.AlignCenter)
                            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                            self.tbl_KStuff.setItem(valWdg, 10, item)
                            if noneFlag == False:
                                item.setToolTip(
                                    "<font color=black>%s</font>" % str(temp2_row[12]).replace("\n", "<br/>"))
                            else:
                                noneFlag = False


                            valWdg += 1
            if checkParentTransfer == False:
                cur.execute("SELECT ChildID, ChildNameObj, ConnectionID FROM treechildobjtbl")
                childList_rows = cur.fetchall()
                # print(str(GlobalValues.TreeParentName))
                for temp1_child_row in childList_rows:
                    ChildvalWdg = 0
                    if str(temp1_child_row[1]) == str(currItemText):
                        currChildID = str(temp1_child_row[0])
                        for temp2_child_row in Parent_rows:
                            if (str(temp2_child_row[11]) == currChildID and str(temp2_child_row[8]) == str(GlobalValues.TreeParentName)):

                                self.tbl_KStuff.insertRow(ChildvalWdg)

                                item = QtWidgets.QTableWidgetItem(str(temp2_child_row[0]))
                                item.setTextAlignment(Qt.AlignCenter)
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tbl_KStuff.setItem(ChildvalWdg, 0, item)

                                item = QtWidgets.QTableWidgetItem(str(temp2_child_row[1]))
                                item.setTextAlignment(Qt.AlignCenter)
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tbl_KStuff.setItem(ChildvalWdg, 1, item)

                                item = QtWidgets.QTableWidgetItem(str(temp2_child_row[2]))
                                item.setTextAlignment(Qt.AlignCenter)
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tbl_KStuff.setItem(ChildvalWdg, 2, item)

                                item = QtWidgets.QTableWidgetItem(str(temp2_child_row[3]))
                                item.setTextAlignment(Qt.AlignCenter)
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tbl_KStuff.setItem(ChildvalWdg, 3, item)

                                item = QtWidgets.QTableWidgetItem(str(temp2_child_row[4]))
                                item.setTextAlignment(Qt.AlignCenter)
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tbl_KStuff.setItem(ChildvalWdg, 4, item)

                                item = QtWidgets.QTableWidgetItem(str(temp2_child_row[5]))
                                item.setTextAlignment(Qt.AlignCenter)
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tbl_KStuff.setItem(ChildvalWdg, 5, item)

                                item = QtWidgets.QTableWidgetItem(str(temp2_child_row[6]))
                                item.setTextAlignment(Qt.AlignCenter)
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tbl_KStuff.setItem(ChildvalWdg, 6, item)

                                item = QtWidgets.QTableWidgetItem(str(temp2_child_row[7]))
                                item.setTextAlignment(Qt.AlignCenter)
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tbl_KStuff.setItem(ChildvalWdg, 7, item)

                                item = QtWidgets.QTableWidgetItem(str(temp2_child_row[8]))
                                item.setTextAlignment(Qt.AlignCenter)
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tbl_KStuff.setItem(ChildvalWdg, 8, item)

                                item = QtWidgets.QTableWidgetItem(str(temp2_child_row[9]))
                                item.setTextAlignment(Qt.AlignCenter)
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tbl_KStuff.setItem(ChildvalWdg, 9, item)

                                noneFlag = False
                                if temp2_child_row[12] == None:
                                    item = QtWidgets.QTableWidgetItem("")
                                    noneFlag = True
                                else:
                                    item = QtWidgets.QTableWidgetItem(temp2_child_row[12])
                                item.setTextAlignment(Qt.AlignCenter)
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tbl_KStuff.setItem(ChildvalWdg, 10, item)
                                if noneFlag == False:
                                    item.setToolTip(
                                        "<font color=black>%s</font>" % str(temp2_child_row[12]).replace("\n", "<br/>"))
                                else:
                                    noneFlag = False



                                ChildvalWdg += 1
        GlobalValues.checkLoad = False

        if GlobalValues.checkTreeStatesUpdate == True:
            GlobalValues.checkTreeStatesUpdate = False
            self.btn_RefreshTreeStates.click()

        GlobalValues.CHECK_KTBL_UPDATE = False

    def DeleteKRecord(self):
        # print(str(self.tbl_OStuff.currentIndex().row()))
        checkNull = -1
        if (int(self.tbl_KStuff.currentIndex().row()) == checkNull):
            GlobalValues.MesText = "Выберите оборудованиeе для удаления!"
            self.openMesBox()
        else:
            self.openPanelDelReason()
            if GlobalValues.checkDelComment == True:
                GlobalValues.checkDelComment = False
                if GlobalValues.MesResult == True:
                    GlobalValues.MesResult = False
                    con = pymysql.connect(host=str(GlobalValues.SqlHostname),
                                          port=int(GlobalValues.SqlPort),
                                          user=str(GlobalValues.SqlUserName),
                                          passwd=str(GlobalValues.SqlPwd),
                                          db=str(GlobalValues.SqlDBName))
                    KID = int(GlobalValues.KTblCurrID)
                    #print(str(KID))
                    with con:

                        cur = con.cursor()

                        query = ("DELETE FROM kstuff where id = (%s)")
                        cur.execute(query,(KID))

                        con.commit()
                        GlobalValues.Hystory(GlobalValues.UserName, GlobalValues.KTblCurrTypeObj,
                                             GlobalValues.KTblCurrSerialNumObj,
                                             "Удаление. Причина: " + str(GlobalValues.DelReasonComment))

                    if GlobalValues.checkGlobalviewMode == False:
                        GlobalValues.checkTreeUpdateEvent = True
                    else:
                        GlobalValues.checkKTblUpdateEvent = True
                    GlobalValues.checkTreeStatesUpdate = True


    def cell2_was_clicked(self):
        row = self.tbl_KStuff.currentIndex().row()
        col = self.tbl_KStuff.currentIndex().column()
        GlobalValues.KTblCurrTransferDateObj = ""
        itemID = self.tbl_KStuff.item(row, 0).text()
        typeObj = self.tbl_KStuff.item(row, 1).text()
        modelObj = self.tbl_KStuff.item(row, 2).text()
        serialNumObj = self.tbl_KStuff.item(row, 3).text()
        makerObj = self.tbl_KStuff.item(row, 4).text()
        distributorObj = self.tbl_KStuff.item(row, 5).text()
        inputDateObj = self.tbl_KStuff.item(row, 6).text()
        targetObj = self.tbl_KStuff.item(row, 7).text()
        factObj = self.tbl_KStuff.item(row, 8).text()
        transferDateObj = self.tbl_KStuff.item(row, 9).text()
        comment = self.tbl_KStuff.item(row, 10).text()

        self.KTblItemID = itemID
        GlobalValues.KTblCurrID = self.KTblItemID
        GlobalValues.KTblCurrTypeObj = str(typeObj)
        GlobalValues.KTblCurrModelObj = str(modelObj)
        GlobalValues.KTblCurrSerialNumObj = str(serialNumObj)
        GlobalValues.KTblCurrMakerObj = str(makerObj)
        GlobalValues.KTblCurrDistributorObj = str(distributorObj)
        GlobalValues.KTblCurrInputDateObj = str(inputDateObj)
        GlobalValues.KTblCurrTargetObj = str(targetObj)
        GlobalValues.KTblCurrFactObj = str(factObj)
        GlobalValues.KTblCurrTransferDateObj = str(transferDateObj)
        GlobalValues.KTblCurrRow = str(row)
        GlobalValues.KTblCurrCol = str(col)
        GlobalValues.KTblCurrComment = str(comment)
        GlobalValues.checkKTblWasClicked = True

    def openKTransfer(self):
        if GlobalValues.checkKTblWasClicked == True :
            if GlobalValues.checkGroupKTbl == True:
                temp_lst_selected = []
                lst_selected = []
                selected = self.tbl_KStuff.selectedItems()
                for item in selected:
                    if item.column() == 0:
                        temp_lst_selected.append(item.data(0))
                    if item.column() == 1:
                        temp_lst_selected.append(item.data(0))
                    if item.column() == 3:
                        temp_lst_selected.append(item.data(0))
                    if item.column() == 8:
                        temp_lst_selected.append(item.data(0))
                        lst_selected.append(temp_lst_selected)
                        temp_lst_selected = []

                GlobalValues.ktbl_selected_lst = lst_selected
                self.uiKGroupTransfer = Ui_KGroupTransfer()
                self.uiKGroupTransfer.exec_()
            else:
                self.uiKTransfer = Ui_KTransfer()
                self.uiKTransfer.exec_()
        else:
            GlobalValues.MesText = "Оборудование не выбрано!"
            self.openMesBox()

    def BackToMain2(self):
        self.le_KViewMode.setText("Глобальный")
        self.tree_Obj.clearSelection()
        GlobalValues.checkGlobalviewMode = True
        GlobalValues.checkKTblUpdateEvent = True

    def openPanelEditKStuff(self):
        if GlobalValues.checkKTblWasClicked == True:
            if GlobalValues.checkGroupKTbl == True:
                temp_lst_selected = []
                lst_selected = []
                selected = self.tbl_KStuff.selectedItems()
                for item in selected:
                    if item.column() == 0:
                        temp_lst_selected.append(item.data(0))
                    if item.column() == 1:
                        temp_lst_selected.append(item.data(0))
                    if item.column() == 3:
                        temp_lst_selected.append(item.data(0))
                    if item.column() == 10:
                        temp_lst_selected.append(item.data(0))
                        lst_selected.append(temp_lst_selected)
                        temp_lst_selected = []
                GlobalValues.ktbl_selected_lst = lst_selected
                self.uiKGroupEdit = Ui_KGroupEdit()
                self.uiKGroupEdit.exec_()
            else:
                self.uiEditKStuff = Ui_EditKStuff()
                self.uiEditKStuff.exec_()
        else:
            GlobalValues.MesText = "Оборудование не выбрано!"
            self.openMesBox()

    def SearchKStuff(self):

        if (self.le_KStuff.text() == ""):
            GlobalValues.MesText = "Введите текст для поика!"
            self.openMesBox()
        else:
            self.tbl_KStuff.setRowCount(0)
            curSearchType = self.cb_KStuff.currentText()
            con = pymysql.connect(host=str(GlobalValues.SqlHostname),
                                  port=int(GlobalValues.SqlPort),
                                  user=str(GlobalValues.SqlUserName),
                                  passwd=str(GlobalValues.SqlPwd),
                                  db=str(GlobalValues.SqlDBName))
            curSearchText = '%' + str(self.le_KStuff.text()) + '%'
            # print(curSearchType)
            with con:
                cur = con.cursor()
                if curSearchType == "Серийный номер":
                    cur.execute(
                        "SELECT ID, KType, KModel, KSerialNum, KMaker, KDistributor, KComeTime, KTargetObj, KFactObj, KGoneToObjTime, ParentID, ChildID, KComment FROM kstuff WHERE KSerialNum Like '" + curSearchText + "'")
                if curSearchType == "Поставщик":
                    cur.execute(
                        "SELECT ID, KType, KModel, KSerialNum, KMaker, KDistributor, KComeTime, KTargetObj, KFactObj, KGoneToObjTime, ParentID, ChildID, KComment FROM kstuff WHERE KDistributor Like '" + curSearchText + "'")
                if curSearchType == "Цел.Объект":
                    cur.execute(
                        "SELECT ID, KType, KModel, KSerialNum, KMaker, KDistributor, KComeTime, KTargetObj, KFactObj, KGoneToObjTime, ParentID, ChildID, KComment FROM kstuff WHERE KTargetObj Like '" + curSearchText + "'")

                rows = cur.fetchall()
                valWdg = 0
                self.tbl_KStuff.setRowCount(0)
                for row in rows:
                    self.tbl_KStuff.insertRow(valWdg)

                    item = QtWidgets.QTableWidgetItem(str(row[0]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 0, item)

                    item = QtWidgets.QTableWidgetItem(str(row[1]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 1, item)

                    item = QtWidgets.QTableWidgetItem(str(row[2]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 2, item)

                    item = QtWidgets.QTableWidgetItem(str(row[3]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 3, item)

                    item = QtWidgets.QTableWidgetItem(str(row[4]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 4, item)

                    item = QtWidgets.QTableWidgetItem(str(row[5]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 5, item)

                    item = QtWidgets.QTableWidgetItem(str(row[6]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 6, item)

                    item = QtWidgets.QTableWidgetItem(str(row[7]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 7, item)

                    item = QtWidgets.QTableWidgetItem(str(row[8]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 8, item)

                    item = QtWidgets.QTableWidgetItem(str(row[9]))
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 9, item)

                    noneFlag = False
                    if row[12] == None:
                        item = QtWidgets.QTableWidgetItem("")
                        noneFlag = True
                    else:
                        item = QtWidgets.QTableWidgetItem(row[12])
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                    self.tbl_KStuff.setItem(valWdg, 10, item)
                    if noneFlag == False:
                        item.setToolTip("<font color=black>%s</font>" % str(row[12]).replace("\n", "<br/>"))
                    else:
                        noneFlag = False

                    valWdg += 1
                cur.close()
            con.close()

    def KTblMode(self):
        if self.chB_KGroup.isChecked() == True:
            GlobalValues.checkGroupKTbl = True
            self.tbl_KStuff.clearSelection()
            self.tbl_KStuff.setSelectionMode(QtWidgets.QAbstractItemView.MultiSelection)
        else:
            GlobalValues.checkGroupKTbl = False
            self.tbl_KStuff.clearSelection()
            self.tbl_KStuff.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)

    def getStatKTbl(self):
        GlobalValues.CurrStatArray = self.getTblStat(self.tbl_KStuff)
        self.openPanelStat()
    # общее
    def openPanelJournal(self):
        self.uiJournal = Ui_Journal()
        self.uiJournal.exec_()

    def closeEvent(self, event):
        GlobalValues.StopAll = True
        event.accept()

    def eventFilter(self, source, event):

        if (event.type() == QtCore.QEvent.Wheel and source is self.tbl_OStuff.viewport()):
            # print("OSCROLL")
            self.vScrl_OTbl.setValue(self.tbl_OStuff.verticalScrollBar().value())

        if (event.type() == QtCore.QEvent.Wheel and source is self.tbl_KStuff.viewport()):
            # print('KKKKKKKScroll!!!!')
            self.vScrl_KTbl_2.setValue(self.tbl_KStuff.verticalScrollBar().value())

        if(event.type() == QtCore.QEvent.MouseButtonPress and
           event.buttons() == QtCore.Qt.RightButton and
           source is self.tbl_OStuff.viewport()):
            GlobalValues.checkWichTblWasClicked = False
            item = self.tbl_OStuff.itemAt(event.pos())
            # print('Global Pos:', event.globalPos())
            if item is not None:
                # print('Table Item:', item.row(), item.column())
                GlobalValues.checkOTblWasClicked = True
                GlobalValues.OTblCurrID = self.tbl_OStuff.item(item.row(), 0).text()
                GlobalValues.OTblCurrTypeObj = self.tbl_OStuff.item(item.row(), 1).text()
                GlobalValues.OTblCurrModelObj = self.tbl_OStuff.item(item.row(), 2).text()
                GlobalValues.OTblCurrSerialNumObj = self.tbl_OStuff.item(item.row(), 3).text()
                GlobalValues.OTblCurrMakerObj = self.tbl_OStuff.item(item.row(), 4).text()
                GlobalValues.OTblCurrDistributorObj = self.tbl_OStuff.item(item.row(), 5).text()
                GlobalValues.OTblCurrInputDateObj = self.tbl_OStuff.item(item.row(), 6).text()
                GlobalValues.OTblCurrTargetObj = self.tbl_OStuff.item(item.row(), 7).text()
                GlobalValues.OTblCurrFactObj = self.tbl_OStuff.item(item.row(), 8).text()
                GlobalValues.OTblCurrTransferDateObj = self.tbl_OStuff.item(item.row(), 9).text()
                GlobalValues.OTblCurrComment = self.tbl_OStuff.item(item.row(), 10).text()

                #создание контекста
                self.menu = QMenu(self)
                self.menu.setStyleSheet("""
                    QMenu {border: 1px inset grey; background-color: #fff; color: #000; padding: -1; padding-left: -6;}
                    QMenu:selected {background-color: #ddf; color: #000;}
                """)
                self.menu.setFixedWidth(105)

                self.menu_actionBtn1 = QAction('История', self)
                self.menu_actionBtn1.setData('option1')
                self.menu_actionBtn1.triggered.connect(self.openHystory)
                self.menu.addAction(self.menu_actionBtn1)

                if GlobalValues.AccesGroup == 'admin':
                    self.menu_actionBtn2 = QAction('Редактировать', self)
                    self.menu_actionBtn2.setData('option2')
                    self.menu_actionBtn2.triggered.connect(self.openPanelEditOStuff)
                    self.menu.addAction(self.menu_actionBtn2)

                    self.menu_actionBtn3 = QAction('Переместить', self)
                    self.menu_actionBtn3.setData('option3')
                    self.menu_actionBtn3.triggered.connect(self.openOTransfer)
                    self.menu.addAction(self.menu_actionBtn3)

                    self.menu_actionBtn4 = QAction('Удалить', self)
                    self.menu_actionBtn4.setData('option4')
                    self.menu_actionBtn4.triggered.connect(self.DeleteORecord)
                    self.menu.addAction(self.menu_actionBtn4)

        if (event.type() == QtCore.QEvent.MouseButtonPress and
                event.buttons() == QtCore.Qt.RightButton and
                source is self.tbl_KStuff.viewport()):
            GlobalValues.checkWichTblWasClicked = True
            item = self.tbl_KStuff.itemAt(event.pos())
            if item is not None:
                GlobalValues.checkKTblWasClicked = True
                GlobalValues.KTblCurrID = self.tbl_KStuff.item(item.row(), 0).text()
                GlobalValues.KTblCurrTypeObj = self.tbl_KStuff.item(item.row(), 1).text()
                GlobalValues.KTblCurrModelObj = self.tbl_KStuff.item(item.row(), 2).text()
                GlobalValues.KTblCurrSerialNumObj = self.tbl_KStuff.item(item.row(), 3).text()
                GlobalValues.KTblCurrMakerObj = self.tbl_KStuff.item(item.row(), 4).text()
                GlobalValues.KTblCurrDistributorObj = self.tbl_KStuff.item(item.row(), 5).text()
                GlobalValues.KTblCurrInputDateObj = self.tbl_KStuff.item(item.row(), 6).text()
                GlobalValues.KTblCurrTargetObj = self.tbl_KStuff.item(item.row(), 7).text()
                GlobalValues.KTblCurrFactObj = self.tbl_KStuff.item(item.row(), 8).text()
                GlobalValues.KTblCurrTransferDateObj = self.tbl_KStuff.item(item.row(), 9).text()
                GlobalValues.KTblCurrComment = self.tbl_KStuff.item(item.row(), 10).text()
                #создание контекста
                self.menu = QMenu(self)
                self.menu.setStyleSheet("""
                    QMenu {border: 1px inset grey; background-color: #fff; color: #000; padding: -1; padding-left: -6;}
                    QMenu:selected {background-color: #ddf; color: #000;}
                """)
                self.menu.setFixedWidth(105)

                self.menu_actionBtn1 = QAction('История', self)
                self.menu_actionBtn1.setData('option1')
                self.menu_actionBtn1.triggered.connect(self.openHystory)
                self.menu.addAction(self.menu_actionBtn1)
                if GlobalValues.AccesGroup == 'admin':
                    self.menu_actionBtn2 = QAction('Редактировать', self)
                    self.menu_actionBtn2.setData('option2')
                    self.menu_actionBtn2.triggered.connect(self.openPanelEditKStuff)
                    self.menu.addAction(self.menu_actionBtn2)

                    self.menu_actionBtn3 = QAction('Переместить', self)
                    self.menu_actionBtn3.setData('option3')
                    self.menu_actionBtn3.triggered.connect(self.openKTransfer)
                    self.menu.addAction(self.menu_actionBtn3)


                    self.menu_actionBtn4 = QAction('Удалить', self)
                    self.menu_actionBtn4.setData('option4')
                    self.menu_actionBtn4.triggered.connect(self.DeleteKRecord)
                    self.menu.addAction(self.menu_actionBtn4)

        if (event.type() == QtCore.QEvent.MouseButtonPress and
                event.buttons() == QtCore.Qt.RightButton and
                source is self.tree_Obj.viewport()):
            item = self.tree_Obj.itemAt(event.pos())
            GlobalValues.treeParentNameByRightClick = item.text(0)


        return super(Ui_MainDB, self).eventFilter(source, event)

    def generateMenu(self, pos):
        # print("pos======", pos)
        if GlobalValues.checkWichTblWasClicked == False:
            self.menu.exec(self.tbl_OStuff.mapToGlobal(pos))
        else:
            self.menu.exec(self.tbl_KStuff.mapToGlobal(pos))

    def generateTreeMenu(self, pos):
        self.menuTree.exec(self.tree_Obj.mapToGlobal(pos))

    def generateBtnMenu(self, pos):
        self.menuSpec.exec(self.btn_Spec.mapToGlobal(pos))

    def TreeClickEvent(self):
        GlobalValues.checkLoad = True
        # print("TreeClicked!!!")

        # self.Load()
        item = self.tree_Obj.currentItem()
        if item.parent() != None:
            item1 = item.parent()
            GlobalValues.TreeParentName = item1.text(0)
            ParamValue = str(item1.text(0))
        else:
            GlobalValues.TreeParentName = item.text(0)
            ParamValue = str(item.text(0))
        ParamName = "treeParentName"
        self.setglobalValueToBD(ParamValue, ParamName)



        # GlobalValues.TreeParentName =  str(item.text(0))
        # print(str(GlobalValues.TreeParentName))

        self.le_OViewMode.setText("Объект")
        self.le_KViewMode.setText("Объект")
        GlobalValues.checkGlobalviewMode = False
        GlobalValues.checkTreeUpdateEvent = True
        # self.setOTblByTreeClick()
        # self.setKTblByTreeClick()
        GlobalValues.checkTreeFocusInTransfer = True

    def setglobalValueToBD(self, value, name):
        con = pymysql.connect(host=str(GlobalValues.SqlHostname),
                              port=int(GlobalValues.SqlPort),
                              user=str(GlobalValues.SqlUserName),
                              passwd=str(GlobalValues.SqlPwd),
                              db=str(GlobalValues.SqlDBName))
        with con:
            cur = con.cursor()
            query = ("UPDATE globaltbl SET paramValue = (%s) WHERE ParamName = (%s)")
            cur.execute(query, (str(value), str(name)))
            cur.close()
        con.close()

    def searchSpecInSource(self, sourceName):
        globcheck = False
        localcheck = False
        check = False
        locsourceLst = os.listdir(GlobalValues.localSpecSource)
        for elem in locsourceLst:
            if elem.find(str(sourceName)) != -1:
                localcheck = True
        if localcheck == True:
            check = True
        else:
            globsourceLst = os.listdir(GlobalValues.globalSpecSource)
            for elem in globsourceLst:
                if elem.find(str(sourceName)) != -1:
                    globcheck = True
        if globcheck == True:
            fileparthsource = GlobalValues.globalSpecSource + sourceName + '.pdf'
            fileparthsource = fileparthsource.replace('/', '\\')
            fileparthdist = GlobalValues.localSpecSource + sourceName + '.pdf'
            fileparthdist = fileparthdist.replace('/', '\\')
            copy(fileparthsource, fileparthdist)
            time.sleep(0.5)
            check = True
        else:
            if localcheck == True:
                check = True
            else:
                check = False

        return check

    def openPanelDelReason(self):
        self.uiDelReason = Ui_DelReason()
        self.uiDelReason.exec_()

    def openPanelStat(self):
        self.uiPanelStat = Ui_Stat()
        self.uiPanelStat.exec_()

    def getTblStat(self, currtbl):
        arr = []
        for i in range(currtbl.rowCount()):
            parr = []
            temp_typeObj = currtbl.item(i, 1).text()
            temp_ModelObj = currtbl.item(i, 2).text()
            parr.append(str(temp_typeObj))
            parr.append(str(temp_ModelObj))
            arr.append(parr)
        return arr

    def openDocumentation(self):
        docName = GlobalValues.TreeParentName
        docPath = str(GlobalValues.globalDocSource) + docName + '/'
        docPath = docPath.replace('/', '\\')
        # print(docPath)
        checkpath = os.path.exists(docPath)
        if checkpath == True:
            # docPath = docPath.replace('\\', '/')
            os.system(r"explorer.exe " + str(docPath))
        else:
            os.makedirs(docPath)
            os.system(r"explorer.exe " + str(docPath))

    def openPhotoArchive(self):
        docName = GlobalValues.TreeParentName
        docPath = str(GlobalValues.globalPhotosource) + docName + '/'
        docPath = docPath.replace('/', '\\')
        # print(docPath)
        checkpath = os.path.exists(docPath)
        if checkpath == True:
            # docPath = docPath.replace('\\', '/')
            os.system(r"explorer.exe " + str(docPath))
        else:
            os.makedirs(docPath)
            os.system(r"explorer.exe " + str(docPath))

    def openTranSource(self):
        docName = GlobalValues.TreeParentName
        docPath = str(GlobalValues.globalTranSource) + docName + '/'
        docPath = docPath.replace('/', '\\')
        # print(docPath)
        checkpath = os.path.exists(docPath)
        if checkpath == True:
            # docPath = docPath.replace('\\', '/')
            os.system(r"explorer.exe " + str(docPath))
        else:
            os.makedirs(docPath)
            os.system(r"explorer.exe " + str(docPath))

    def openNetworkConf(self):
        docName = GlobalValues.TreeParentName
        docPath = str(GlobalValues.globalNetConfSource) + docName + '/'
        docPath = docPath.replace('/', '\\')
        # print(docPath)
        checkpath = os.path.exists(docPath)
        if checkpath == True:
            # docPath = docPath.replace('\\', '/')
            os.system(r"explorer.exe " + str(docPath))
        else:
            os.makedirs(docPath)
            os.system(r"explorer.exe " + str(docPath))

    def thOSPEC(self):
        checkSpec = self.searchSpecInSource(GlobalValues.TreeParentName)
        if checkSpec == False:
            GlobalValues.MesText = "Файлы не найдены! Загрузить?"
            self.openMesBox()
            if GlobalValues.MesResult == True:
                self.loadSpec()
                if GlobalValues.checkSpeckOpenBoxCancel != True:
                    self.btn_Spec.click()
        else:
            # th = threading.Thread(target=self.oSpecFunc)
            # th.start()
            oSpecFun()

    def loadSpec(self):
        GlobalValues.SpecPath = ''
        GlobalValues.checkSpeckPath = False
        GlobalValues.checkSpeckOpenBoxCancel = False
        currSpecName = GlobalValues.TreeParentName
        fileparthsource = easygui.fileopenbox( default="./data/*.pdf" )
        # self.thLoadBox()
        # while GlobalValues.checkSpeckPath == False:
        #     print("waiting for SpecPath...")
        #     if (GlobalValues.SpecPath != '') or (GlobalValues.checkSpeckOpenBoxCancel == True):
        #         GlobalValues.checkSpeckPath = True

        # if (GlobalValues.checkSpeckPath == True) and (GlobalValues.checkSpeckOpenBoxCancel == False):
        if str(fileparthsource) != "None":
            fileparthdist = GlobalValues.globalSpecSource + currSpecName + '.pdf'
            fileparthsource = fileparthsource.replace('/', '\\')
            fileparthdist = fileparthdist.replace('/', '\\')
            copy(fileparthsource, fileparthdist)
            fileparthdist = GlobalValues.localSpecSource + currSpecName + '.pdf'
            fileparthdist = fileparthdist.replace('/', '\\')
            copy(fileparthsource, fileparthdist)
        else:
            GlobalValues.checkSpeckOpenBoxCancel = True

    def reloadSpec(self):
        checkSpec = self.searchSpecInSource(GlobalValues.TreeParentName)
        if checkSpec == False:
            GlobalValues.MesText = "Файлы не найдены! Загрузить?"
            self.openMesBox()
            if GlobalValues.MesResult == True:
                self.loadSpec()
                if GlobalValues.checkSpeckOpenBoxCancel != True:
                    self.btn_Spec.click()
        else:
            self.loadSpec()
            if GlobalValues.checkSpeckOpenBoxCancel != True:
                self.btn_Spec.click()
            # print("1111")

    def openGlobalStat(self):
        self.uiGlobalStat = Ui_GlobalStat()
        self.uiGlobalStat.exec_()

    def renameTreeObj(self):
        oldObjName = GlobalValues.treeParentNameByRightClick

        GlobalValues.DialogHeadText = "Новое имя: "
        GlobalValues.DialogText = GlobalValues.treeParentNameByRightClick
        self.openDialogBox()
        if GlobalValues.DialogRes == True:
            newObjName = GlobalValues.DialogText
            #ренейм в базе объектов
            con = pymysql.connect(host=str(GlobalValues.SqlHostname),
                                  port=int(GlobalValues.SqlPort),
                                  user=str(GlobalValues.SqlUserName),
                                  passwd=str(GlobalValues.SqlPwd),
                                  db=str(GlobalValues.SqlDBName))
            with con:
                cur = con.cursor()
                query = ("UPDATE treeobjtbl SET ParentObjName = (%s) WHERE ParentObjName = (%s)")
                cur.execute(query, (newObjName, oldObjName))

                query = ("UPDATE ostuff SET OTargetObj = (%s) WHERE OTargetObj = (%s)")
                cur.execute(query, (newObjName, oldObjName))
                query = ("UPDATE ostuff SET OFactObj = (%s) WHERE OFactObj = (%s)")
                cur.execute(query, (newObjName, oldObjName))
                query = ("UPDATE ostuff SET OInputObj = (%s) WHERE OInputObj = (%s)")
                cur.execute(query, (newObjName, oldObjName))

                query = ("UPDATE kstuff SET KTargetObj = (%s) WHERE KTargetObj = (%s)")
                cur.execute(query, (newObjName, oldObjName))
                query = ("UPDATE kstuff SET KFactObj = (%s) WHERE KFactObj = (%s)")
                cur.execute(query, (newObjName, oldObjName))
                query = ("UPDATE kstuff SET KInputObj = (%s) WHERE KInputObj = (%s)")
                cur.execute(query, (newObjName, oldObjName))
            #возвращение фокуса на элемент
            self.treeUpdate()
            count = self.tree_count_tems()
            i = 0
            for i in range(count):
                self.tree_Obj.setCurrentItem(self.tree_Obj.topLevelItem(i))
                item = self.tree_Obj.currentItem()
                itemText = item.text(0)
                if itemText == newObjName:
                    break
                i+=1
            #ренейм каталогов
            pathDoc = GlobalValues.globalDocSource + oldObjName + '/'
            pathPhoto = GlobalValues.globalPhotosource + oldObjName + '/'
            pathSpec = GlobalValues.globalSpecSource + oldObjName + '.pdf'
            pathNet = GlobalValues.globalNetConfSource + oldObjName + '/'
            pathTran = GlobalValues.globalTranSource + oldObjName + '/'

            oldPathDoc = pathDoc.replace('/', '\\')
            oldPathPhoto = pathPhoto.replace('/', '\\')
            oldPathSpec = pathSpec.replace('/', '\\')
            oldPathNet = pathNet.replace('/', '\\')
            oldPathTran = pathTran.replace('/', '\\')

            pathDoc = GlobalValues.globalDocSource + newObjName + '/'
            pathPhoto = GlobalValues.globalPhotosource + newObjName + '/'
            pathSpec = GlobalValues.globalSpecSource + newObjName + '.pdf'
            pathNet = GlobalValues.globalNetConfSource + newObjName + '/'
            pathTran = GlobalValues.globalTranSource + newObjName + '/'

            newPathDoc = pathDoc.replace('/', '\\')
            newPathPhoto = pathPhoto.replace('/', '\\')
            newPathSpec = pathSpec.replace('/', '\\')
            newPathNet = pathNet.replace('/', '\\')
            newPathTran = pathTran.replace('/', '\\')



            checkpath = os.path.exists(oldPathDoc)
            if checkpath == True:
                GlobalValues.TreeParentName = newObjName
                os.rename(oldPathDoc, newPathDoc)

            checkpath = os.path.exists(oldPathPhoto)
            if checkpath == True:
                GlobalValues.TreeParentName = newObjName
                os.rename(oldPathPhoto, newPathPhoto)

            checkpath = os.path.exists(oldPathSpec)
            if checkpath == True:
                GlobalValues.TreeParentName = newObjName
                os.rename(oldPathSpec, newPathSpec)

            checkpath = os.path.exists(oldPathNet)
            if checkpath == True:
                GlobalValues.TreeParentName = newObjName
                os.rename(oldPathNet, newPathNet)

            checkpath = os.path.exists(oldPathTran)
            if checkpath == True:
                GlobalValues.TreeParentName = newObjName
                os.rename(oldPathTran, newPathTran)


            GlobalValues.checkTreeStatesUpdate = True
            if GlobalValues.checkGlobalviewMode == True:
                GlobalValues.checkOTblUpdateEvent = True
                GlobalValues.checkKTblUpdateEvent = True
            else:
                GlobalValues.checkTreeUpdateEvent = True

    def tree_count_tems(self):
        count = 0
        iterator = QtWidgets.QTreeWidgetItemIterator(self.tree_Obj)  # pass your treewidget as arg
        while iterator.value():
            item = iterator.value()

            if item.parent():
                if item.parent().isExpanded():
                    count += 1
            else:
                # root item
                count += 1
            iterator += 1

        return count

    def getTreeElemStat(self, db_con, objName):
        pType = GlobalValues.STblCurrTypeObj
        pModel = GlobalValues.STblCurrModelObj
        pTargetObj = objName

        con = db_con

        cur = con.cursor()
        cur.execute(
            "SELECT OTypeName FROM otypetbl WHERE OTypeName Like '" + str(pType) + "'")
        orows = cur.fetchall()

        if str(len(orows)) == '1':
            checkObj = False
        else:
            checkObj = True
        # print(checkObj)

        if checkObj == False:
            cur.execute(
                "SELECT OType, OModel, OTargetObj, OFactObj FROM ostuff")
            StuffRows = cur.fetchall()

        else:
            cur.execute(
                "SELECT KType, KModel, KTargetObj, KFactObj FROM kstuff")
            StuffRows = cur.fetchall()

        pFactNum = 0

        for row in StuffRows:
            if (str(row[2]) == pTargetObj) or (str(row[3]) == pTargetObj):
                if str(row[0]) == pType:
                    if str(row[1]) == pModel:
                        if (str(row[3]) == 'Склад Офис') or (str(row[3]) == pTargetObj):
                            pFactNum +=1


        pFactNum = str(pFactNum)
        return pFactNum

    def checkObjSomeBody(self, con, itemInd):
        cur = con.cursor()
        cur.execute(
            "SELECT ParentObjName FROM treeobjtbl WHERE ConnectionID Like '" + str(itemInd) + "'")
        rows = cur.fetchall()

        for row in rows:
            curObjName = row[0]

        cur.execute(
            "SELECT OTargetObj, OFactObj FROM ostuff WHERE OTargetObj Like '" + str(curObjName) + "'")
        ostuff_target_rows = cur.fetchall()

        cur.execute(
            "SELECT OTargetObj, OFactObj FROM ostuff WHERE OFactObj Like '" + str(curObjName) + "'")
        ostuff_fact_rows = cur.fetchall()

        cur.execute(
            "SELECT KTargetObj, KFactObj FROM kstuff WHERE KTargetObj Like '" + str(curObjName) + "'")
        kstuff_target_rows = cur.fetchall()

        cur.execute(
            "SELECT KTargetObj, KFactObj FROM kstuff WHERE KFactObj Like '" + str(curObjName) + "'")
        kstuff_fact_rows = cur.fetchall()


        countCurr = 0
        countNotCurr = 0
        for row in ostuff_target_rows:
            if row[1] == 'Склад Офис':
                countCurr +=1
            if row[0] == row[1]:
                countCurr +=1
            if (row[1] != row[0]) and (row[1] != 'Склад Офис'):
                countNotCurr +=1

        for row in ostuff_fact_rows:
            if row[1] != row[0]:
                countCurr+=1

        for row in kstuff_target_rows:
            if row[1] == 'Склад Офис':
                countCurr +=1
            if row[0] == row[1]:
                countCurr +=1
            if (row[1] != row[0]) and (row[1] != 'Склад Офис'):
                countNotCurr +=1

        for row in kstuff_fact_rows:
            if row[1] != row[0]:
                countCurr+=1

        return countCurr, countNotCurr, curObjName

    def getTreeObjStat(self, itemInd):
        curConID = 0
        checkObj = True
        checkNull = True
        RES = '2'
        locHostName = str(GlobalValues.SqlHostname)
        locPort = int(GlobalValues.SqlPort)
        locUserName = str(GlobalValues.SqlUserName)
        locPwd = str(GlobalValues.SqlPwd)
        locDBName = str(GlobalValues.SqlDBName)
        con = pymysql.connect(host=locHostName,
                              port=locPort,
                              user=locUserName,
                              passwd=locPwd,
                              db=locDBName)

        cur = con.cursor()
        # cur.execute(
        #     "SELECT ConnectionID FROM treeobjtbl WHERE ParentObjName Like '" + str(GlobalValues.TreeParentName) + "'")
        # rows = cur.fetchall()
        #
        #
        # for row in rows:
        #     curConID = row[0]
        # print('COnnectID: ', str(curConID))

        cur.execute(
            "SELECT TypeObj, ModelObj, Num, ConID FROM globstattbl WHERE ConID Like '" + str(itemInd) + "'")
        statCellRows = cur.fetchall()
        # print(statCellRows)
        statCount = 0
        specNumCount = 0

        currStuffCount,notCurrStuffCount,curParentName = self.checkObjSomeBody(con, itemInd)


        if str(len(statCellRows)) == '0':
            if (currStuffCount == 0) and (notCurrStuffCount == 0):
                RES = '0'
            else:
                RES = '1'

        else:
            for row in statCellRows:
                GlobalValues.STblCurrTypeObj = row[0]
                GlobalValues.STblCurrModelObj = row[1]
                GlobalValues.STblCurrNumBySpec = row[2]

                factNum = self.getTreeElemStat(con, str(curParentName))
                statCount = statCount + int(factNum)
                specNumCount = specNumCount + int(GlobalValues.STblCurrNumBySpec)
                if str(row[2]) != str(factNum):
                    checkObj = False

                if str(factNum) != '0':
                    checkNull = False

            if checkObj == True:
                if statCount == currStuffCount:
                    RES = '2'
                else:
                    RES = '1'
            else:
                if (currStuffCount != 0) or (notCurrStuffCount != 0):
                    RES = '1'
                else:
                    RES = '0'
        # print('ObjName: ', str(GlobalValues.TreeParentName))
        # print('currStuffCount: ', str(currStuffCount))
        # print('statCount: ', str(statCount))
        # print('notCurrStuffCount: ', str(notCurrStuffCount))


        return RES

    def thTreeColorStateUpdate(self):
        th = threading.Thread(target=self.treeColorStateUpdate)
        th.start()

    def treeColorStateUpdate(self):
        try:
            treecount = self.tree_Obj.topLevelItemCount()
            count = treecount + 3
            i = 0
            for i in range(0, count, 1):
                GlobalValues.TH_OBJ_COUNTER = i
                if i > 3:
                    th = threading.Thread(target= self.getObjState)
                    th.start()
                i+=1


            # while treecount != GlobalValues.TH_OBJ_COUNTER:
            #     self.tree_Obj.raise_()
            #     GlobalValues.TH_OBJ_COUNTER = 0
        except Exception as ex:
            print('treeStates thread FAIL!')

    def getObjState(self):
        try:
            ind = GlobalValues.TH_OBJ_COUNTER
            curInd = ind - 3
            i = ind
            checkObj = self.getTreeObjStat(i)
            if checkObj == '2':
                self.tree_Obj.topLevelItem(curInd).setBackground(0, QtGui.QBrush(self.green_gradient))
            if checkObj == '1':
                self.tree_Obj.topLevelItem(curInd).setBackground(0, QtGui.QBrush(self.yellow_gradient))
            if checkObj == '0':
                self.tree_Obj.topLevelItem(curInd).setBackground(0, QtGui.QBrush(self.red_gradient))
            self.MainFrame.raise_()
            self.MainFrame.lower()
                # print(curInd)
        except:
            print('TH OBJ STATE WAS FAIL!!! (' + str(curInd) + ')')
            self.getObjState(ind)

    def openDialogBox(self):
        self.uiDialogBox = Ui_DialogBox()
        self.uiDialogBox.exec_()

    def openExelDataTable(self):
        path = GlobalValues.globalExelDataTableSourse
        path = path.replace('/', '\\')
        os.startfile(str(path))

    def openInfo(self):
        self.uiInfo = Ui_Info()
        self.uiInfo.exec_()

    def as_text(self, value):
        if value is None:
            return ""
        return str(value)

    def makeExel(self):
        curObjName = GlobalValues.treeParentNameByRightClick
        locHostName = str(GlobalValues.SqlHostname)
        locPort = int(GlobalValues.SqlPort)
        locUserName = str(GlobalValues.SqlUserName)
        locPwd = str(GlobalValues.SqlPwd)
        locDBName = str(GlobalValues.SqlDBName)
        con = pymysql.connect(host=locHostName,
                              port=locPort,
                              user=locUserName,
                              passwd=locPwd,
                              db=locDBName)

        cur = con.cursor()

        cur.execute(
            "SELECT OType, OModel, OSerialNum, OComment, ChildID FROM ostuff WHERE OFactObj Like '" + str(curObjName) + "'")
        oFactRows = cur.fetchall()
        time.sleep(0.2)

        cur.execute(
            "SELECT OType, OModel, OSerialNum, OComment, OFactObj, ChildID FROM ostuff WHERE OTargetObj Like '" + str(curObjName) + "'")
        oTargetRows = cur.fetchall()
        time.sleep(0.2)

        cur.execute(
            "SELECT KType, KModel, KSerialNum, KComment, ChildID FROM kstuff WHERE KFactObj Like '" + str(curObjName) + "'")
        kFactRows = cur.fetchall()
        time.sleep(0.2)

        cur.execute(
            "SELECT KType, KModel, KSerialNum, KComment, KFactObj, ChildID FROM kstuff WHERE KTargetObj Like '" + str(curObjName) + "'")
        kTargetRows = cur.fetchall()
        time.sleep(0.2)

        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        # Data can be assigned directly to cells
        ws['A1'] = 'Наименование'
        ws.alignment = Alignment(horizontal='center')
        ws['B1'] = 'Тип, Модель'
        ws['C1'] = 'Ед.Изм.'
        ws['D1'] = 'Шт.'
        ws['E1'] = 'Серийный номер'
        ws['F1'] = 'Расположение'

        for row in oFactRows:
            curChildName = ''
            cur.execute(
                "SELECT ChildNameObj FROM treechildobjtbl WHERE ChildID Like '" + str(row[4]) + "'")
            childNameRows = cur.fetchall()
            for temp_row in childNameRows:
                curChildName = temp_row[0]

            if (row[3] == None) or (row[3] == ''):
                comm = curChildName
            else:
                if curChildName == '':
                    comm = row[3]
                else:
                    comm = str(curChildName) + ' | ' + row[3]
            ws.append([str(row[0]), str(row[1]),  str('шт.'), 1, str(row[2]), comm])

        for row in oTargetRows:
            curChildName = ''
            comm = ''
            if row[4] == 'Склад Офис':
                cur.execute(
                    "SELECT ChildNameObj FROM treechildobjtbl WHERE ChildID Like '" + str(row[5]) + "'")
                childNameRows = cur.fetchall()
                for temp_row in childNameRows:
                    curChildName = temp_row[0]

                if (row[3] == None) or (row[3] == ''):
                    comm = curChildName
                else:
                    if curChildName == '':
                        comm = row[3]
                    else:
                        comm = str(curChildName) + ' | ' + row[3]

                ws.append([str(row[0]), str(row[1]), str('шт.'), 1, str(row[2]), comm])

        for row in kFactRows:
            curChildName = ''
            cur.execute(
                "SELECT ChildNameObj FROM treechildobjtbl WHERE ChildID Like '" + str(row[4]) + "'")
            childNameRows = cur.fetchall()
            for temp_row in childNameRows:
                curChildName = temp_row[0]

            if (row[3] == None) or (row[3] == ''):
                comm = curChildName
            else:
                if curChildName == '':
                    comm = row[3]
                else:
                    comm = str(curChildName) + ' | ' + row[3]

            ws.append([str(row[0]), str(row[1]),  str('шт.'), 1, str(row[2]), comm])

        for row in kTargetRows:
            curChildName = ''
            comm = ''
            if row[4] == 'Склад Офис':
                cur.execute(
                    "SELECT ChildNameObj FROM treechildobjtbl WHERE ChildID Like '" + str(row[5]) + "'")
                childNameRows = cur.fetchall()
                for temp_row in childNameRows:
                    curChildName = temp_row[0]

                if (row[3] == None) or (row[3] == ''):
                    comm = curChildName
                else:
                    if curChildName == '':
                        comm = row[3]
                    else:
                        comm = str(curChildName) + ' | ' + row[3]

                ws.append([str(row[0]), str(row[1]), str('шт.'), 1, str(row[2]), comm])

        data = time.localtime()
        curExName = str('C:\\\\itoDB\dump\dump_') + str(curObjName) + str('_') + str(data[0]) + str(data[1]) + str(data[2]) + str(data[3]) + str(data[4]) + str(data[5]) + '.xlsx'
        curExName = curExName.replace(' ', '_')
        for col in ws.columns:
            for cell in col:
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj

        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(str(cell.value)))
                except IndexError:
                    column_widths.append(len(str(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width




        wb.save(str(curExName))
        os.system('start excel.exe ' + curExName)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    uiMain = Ui_MainDB()
    uiMain.show()
    sys.exit(app.exec_())
